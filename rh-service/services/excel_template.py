"""Gera o .xlsx modelo para a equipe preencher e reenviar via upload."""
import io

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

_COLUNAS = [
    "EMPRESA", "UF", "ALOCAÇÃO REAL", "Nº DA REQUISIÇÃO", "TIPO DO CONTRATO",
    "DATA RECEBIMENTO", "DATA DA APROVAÇÃO DIRETORIA", "TIPO DA VAGA", "CARGO",
    "NÍVEL", "HIERARQUIA", "REQUISITANTE (PRIMEIRO E ÚLTIMO NOME)",
    "STATUS DA VAGA", "ETAPAS DO PROCESSO", "SEÇÃO", "SLA",
    "RESPONSÁVEL PELA VAGA", "JUSTIFICATIVA", "DATA ADMISSÃO OU MOVIMENTAÇÃO",
    "CANDIDATO",
]

# (nome da lista suspensa oficial da aba LISTA -> tabela no banco)
_LISTAS_VALIDADAS = {
    "EMPRESA": ("rh_empresas", "nome"),
    "UF": ("rh_ufs", "sigla"),
    "TIPO DO CONTRATO": ("rh_tipos_contrato", "nome"),
    "TIPO DA VAGA": ("rh_tipos_vaga", "nome"),
    "NÍVEL": ("rh_niveis", "nome"),
    "STATUS DA VAGA": ("rh_status_vaga", "nome"),
    "SEÇÃO": ("rh_secoes", "nome"),
}


def gerar_template(sb) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Controle de Vagas"
    ws.append(_COLUNAS)

    lista_ws = wb.create_sheet("LISTA")
    col_idx = 1
    for header, (tabela, coluna) in _LISTAS_VALIDADAS.items():
        valores = sorted(
            r[coluna] for r in sb.table(tabela).select(coluna).execute().data if r.get(coluna)
        )
        lista_ws.cell(row=1, column=col_idx, value=header)
        for i, v in enumerate(valores, start=2):
            lista_ws.cell(row=i, column=col_idx, value=v)

        letra = lista_ws.cell(row=1, column=col_idx).column_letter
        if valores:
            dv = DataValidation(
                type="list",
                formula1=f"LISTA!${letra}$2:${letra}${len(valores) + 1}",
                allow_blank=True,
            )
            ws.add_data_validation(dv)
            data_col_letra = ws.cell(row=1, column=_COLUNAS.index(header) + 1).column_letter
            dv.add(f"{data_col_letra}2:{data_col_letra}1000")
        col_idx += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
