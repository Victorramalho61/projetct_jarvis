import csv
import io
import logging
import re
import time
import uuid as uuid_mod
from datetime import date, datetime, timedelta, timezone
from typing import Annotated, Any

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from auth import require_role
from db import get_supabase, get_settings
from services.audit import log_action

router = APIRouter(prefix="/api/performance/admin")
_logger = logging.getLogger(__name__)

_RH_ADMIN = ("admin", "rh")

# ── Cache simples em memória (TTL 120 s) ───────────────────────────────────────
# Evita repetir queries idênticas em acessos simultâneos ao dashboard.
# TTL dobrado de 60→120s: reduz carga no PostgreSQL em picos de acesso.
_cache: dict[str, tuple[Any, float]] = {}
_CACHE_TTL = 120.0  # segundos — era 60s, dobrado em hardening 2026-05-28


def _cache_get(key: str) -> Any:
    entry = _cache.get(key)
    if entry and (time.monotonic() - entry[1]) < _CACHE_TTL:
        return entry[0]
    return None


def _cache_set(key: str, value: Any) -> None:
    _cache[key] = (value, time.monotonic())


def _cache_invalidate_prefix(prefix: str) -> None:
    for k in list(_cache.keys()):
        if k.startswith(prefix):
            del _cache[k]

# ── Level mapping ──────────────────────────────────────────────────────────────

_LEVEL_MAP = {
    "gerente": 1,
    "coordenador_supervisor": 2,
    "administrativo": 3,
    "operacional": 3,
}
_PERFIS_VALIDOS = {"gerente", "coordenador_supervisor", "administrativo", "operacional"}
# Sem legado "administrativo_operacional" — todos os colaboradores já foram migrados


def _emp_out(e: dict) -> dict:
    out = dict(e)
    hl = e.get("hierarchy_level") or 3
    perfil = e.get("perfil") or ""
    if hl == 1:
        out["level"] = "gerente"
    elif hl == 2:
        out["level"] = "coordenador_supervisor"
    elif perfil in ("administrativo", "operacional"):
        out["level"] = perfil
    else:
        out["level"] = "administrativo"  # fallback: sem perfil → administrativo
    out["cpf"] = e.get("cpf") or ""
    return out


# ── Helper: current cycle ──────────────────────────────────────────────────────

def _get_current_cycle(db) -> dict | None:
    """Retorna o ciclo mais relevante em uma única query (open > draft > closed)."""
    res = (
        db.table("performance_cycles")
        .select("*")
        .order("created_at", desc=True)
        .execute()
    )
    if not res.data:
        return None
    _priority = {"open": 0, "draft": 1, "closed": 2}
    return min(res.data, key=lambda c: _priority.get(c.get("status"), 99))


# ── Dashboard ──────────────────────────────────────────────────────────────────

@router.get("/dashboard")
def dashboard(
    _: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
    company_id: str | None = None,
    branch_id: str | None = None,
    cycle_id: str | None = None,
) -> dict:
    db = get_supabase()

    if not cycle_id:
        cycle = _get_current_cycle(db)
        cycle_id = cycle["id"] if cycle else None

    if not cycle_id:
        return {
            "total_evaluated": 0, "completion_pct": 0,
            "pending_acknowledgment": 0, "without_evaluation": 0,
            "indicator_averages": [],
            "self_eval_sent": 0, "self_eval_completed": 0, "self_eval_pct": 0,
            "calibrations_count": 0, "calibrations_pct": 0,
        }

    cache_key = f"dashboard:{cycle_id}:{company_id or 'all'}:{branch_id or 'all'}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    emp_q = db.table("performance_employees").select("id").eq("active", True)
    if company_id:
        emp_q = emp_q.eq("company_id", company_id)
    if branch_id:
        emp_q = emp_q.eq("branch_id", branch_id)
    all_employees = emp_q.execute().data
    total_employees = len(all_employees)
    all_emp_ids = {e["id"] for e in all_employees}

    if (company_id or branch_id) and not all_emp_ids:
        result = {
            "total_evaluated": 0, "completion_pct": 0,
            "pending_acknowledgment": 0, "without_evaluation": 0,
            "indicator_averages": [],
            "self_eval_sent": 0, "self_eval_completed": 0, "self_eval_pct": 0,
            "calibrations_count": 0, "calibrations_pct": 0,
        }
        _cache_set(cache_key, result)
        return result

    rev_q = db.table("performance_reviews").select("id,employee_id,status").eq("cycle_id", cycle_id).eq("is_self_evaluation", False)
    if company_id or branch_id:
        rev_q = rev_q.in_("employee_id", list(all_emp_ids))
    reviews = rev_q.execute().data

    completed = [r for r in reviews if r.get("status") in ("completed", "calibrated")]
    reviewed_ids = {r["employee_id"] for r in reviews if r.get("employee_id")}
    without_evaluation = max(0, total_employees - len(reviewed_ids))
    completion_pct = round(len(completed) / total_employees * 100, 1) if total_employees else 0

    review_ids = [r["id"] for r in completed]
    acked_ids: set[str] = set()
    calibrated_ids: set[str] = set()
    if review_ids:
        acks = db.table("performance_review_acknowledgments").select("review_id").in_("review_id", review_ids).execute().data
        acked_ids = {a["review_id"] for a in acks}
        calibs = db.table("performance_calibrations").select("review_id").in_("review_id", review_ids).execute().data
        calibrated_ids = {c["review_id"] for c in calibs}

    pending_ack = len(completed) - len(acked_ids & {r["id"] for r in completed})
    calibrations_count = len(calibrated_ids)
    calibrations_pct = round(calibrations_count / len(completed) * 100, 1) if completed else 0

    # Indicator averages
    indicator_averages: list[dict] = []
    if review_ids:
        scores_raw = (
            db.table("performance_indicator_scores")
            .select("indicator_id,score,performance_indicators(name)")
            .in_("review_id", review_ids)
            .execute()
            .data
        )
        by_ind: dict[str, list[float]] = {}
        ind_names: dict[str, str] = {}
        for s in scores_raw:
            iid = s["indicator_id"]
            ind = s.get("performance_indicators") or {}
            ind_names[iid] = ind.get("name", iid) if isinstance(ind, dict) else iid
            by_ind.setdefault(iid, []).append(float(s["score"]))
        indicator_averages = sorted(
            [{"name": ind_names[k], "avg": round(sum(v) / len(v), 2)} for k, v in by_ind.items()],
            key=lambda x: x["name"],
        )

    # ── Métricas de auto-avaliação ──────────────────────────────────────────────
    _tokens_q = db.table("performance_self_evaluation_tokens").select("employee_id").eq("cycle_id", cycle_id)
    _tokens = _tokens_q.execute().data
    if company_id or branch_id:
        self_eval_sent = sum(1 for t in _tokens if t.get("employee_id") in all_emp_ids)
    else:
        self_eval_sent = len(_tokens)
    self_eval_completed_q = (
        db.table("performance_reviews").select("id")
        .eq("cycle_id", cycle_id).eq("is_self_evaluation", True).eq("status", "completed")
    )
    if company_id or branch_id:
        self_eval_completed_q = self_eval_completed_q.in_("employee_id", list(all_emp_ids))
    self_eval_completed = len(self_eval_completed_q.execute().data)
    self_eval_pct = round(self_eval_completed / total_employees * 100, 1) if total_employees else 0

    result = {
        "total_evaluated": len(completed),
        "completion_pct": completion_pct,
        "pending_acknowledgment": pending_ack,
        "without_evaluation": without_evaluation,
        "indicator_averages": indicator_averages,
        "self_eval_sent": self_eval_sent,
        "self_eval_completed": self_eval_completed,
        "self_eval_pct": self_eval_pct,
        "calibrations_count": calibrations_count,
        "calibrations_pct": calibrations_pct,
    }
    _cache_set(cache_key, result)
    return result


@router.get("/dashboard/pending-evaluators")
def dashboard_pending_evaluators(
    _: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
    cycle_id: str | None = None,
    company_id: str | None = None,
    branch_id: str | None = None,
) -> list[dict]:
    """Managers who still have pending employees to evaluate in the current cycle."""
    db = get_supabase()
    if not cycle_id:
        cycle = _get_current_cycle(db)
        cycle_id = cycle["id"] if cycle else None
    if not cycle_id:
        return []

    # Employees that should be evaluated (L2 + L3)
    emp_q = (
        db.table("performance_employees")
        .select("id,name,cargo,manager_id")
        .in_("hierarchy_level", [2, 3])
        .eq("active", True)
    )
    if company_id:
        emp_q = emp_q.eq("company_id", company_id)
    if branch_id:
        emp_q = emp_q.eq("branch_id", branch_id)
    employees = emp_q.execute().data

    # Employees who already have a completed review in this cycle
    reviewed_ids: set[str] = set()
    if employees:
        all_ids = [e["id"] for e in employees]
        rev_res = (
            db.table("performance_reviews")
            .select("employee_id")
            .eq("cycle_id", cycle_id)
            .in_("status", ["completed", "calibrated"])
            .execute()
            .data
        )
        reviewed_ids = {r["employee_id"] for r in rev_res}

    # Collect pending employees grouped by manager
    pending_by_manager: dict[str, list[dict]] = {}
    for emp in employees:
        if emp["id"] in reviewed_ids:
            continue
        mgr_id = emp.get("manager_id") or "__no_manager__"
        pending_by_manager.setdefault(mgr_id, []).append({"name": emp["name"], "cargo": emp.get("cargo", "")})

    if not pending_by_manager:
        return []

    # Fetch manager details
    mgr_ids = [mid for mid in pending_by_manager if mid != "__no_manager__"]
    managers_map: dict[str, dict] = {}
    if mgr_ids:
        mgrs = db.table("performance_employees").select("id,name,email").in_("id", mgr_ids).execute().data
        managers_map = {m["id"]: m for m in mgrs}

    result = []
    for mgr_id, pending_emps in pending_by_manager.items():
        mgr = managers_map.get(mgr_id, {})
        result.append({
            "manager_name": mgr.get("name", "Sem gestor definido"),
            "manager_email": mgr.get("email", ""),
            "pending_employees": pending_emps,
        })
    result.sort(key=lambda x: x["manager_name"])
    return result


@router.get("/dashboard/pending-self-eval")
def dashboard_pending_self_eval(
    _: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
    cycle_id: str | None = None,
    company_id: str | None = None,
    branch_id: str | None = None,
) -> list[dict]:
    """Colaboradores que ainda não concluíram a auto-avaliação."""
    db = get_supabase()
    if not cycle_id:
        cycle = _get_current_cycle(db)
        cycle_id = cycle["id"] if cycle else None
    if not cycle_id:
        return []

    emp_q = db.table("performance_employees").select("id,name,cargo,hierarchy_level,performance_branches(name),performance_companies(name)").eq("active", True)
    if company_id:
        emp_q = emp_q.eq("company_id", company_id)
    if branch_id:
        emp_q = emp_q.eq("branch_id", branch_id)
    employees = emp_q.execute().data

    completed_self = {
        r["employee_id"]
        for r in db.table("performance_reviews")
        .select("employee_id")
        .eq("cycle_id", cycle_id)
        .eq("is_self_evaluation", True)
        .eq("status", "completed")
        .execute()
        .data
    }

    _PERFIL_LABEL = {"gerente": "Gerente", "coordenador_supervisor": "Coord./Supervisor",
                     "administrativo": "Administrativo", "operacional": "Operacional",
                     }
    _HLEVEL_LABEL = {1: "Gerente", 2: "Coord./Supervisor", 3: "Adm./Operacional"}
    def _level_str(emp_data: dict) -> str:
        p = emp_data.get("perfil") or ""
        return _PERFIL_LABEL.get(p) or _HLEVEL_LABEL.get(emp_data.get("hierarchy_level"), "")

    result = []
    for emp in employees:
        if emp["id"] not in completed_self:
            branch = emp.get("performance_branches") or {}
            company = emp.get("performance_companies") or {}
            result.append({
                "employee_id": emp["id"],
                "employee_name": emp["name"],
                "employee_cargo": emp.get("cargo", ""),
                "hierarchy_level": _level_str(emp),
                "branch_name": branch.get("name", "") if isinstance(branch, dict) else "",
                "company_name": company.get("name", "") if isinstance(company, dict) else "",
            })
    return sorted(result, key=lambda x: x["employee_name"])


@router.get("/dashboard/export")
def dashboard_export(
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
    company_id: str | None = None,
    branch_id: str | None = None,
    cycle_id: str | None = None,
):
    """Exporta dashboard completo em XLSX."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    db = get_supabase()
    if not cycle_id:
        cycle = _get_current_cycle(db)
        cycle_id = cycle["id"] if cycle else None

    # Obter dados do dashboard
    dash = dashboard(current_user, company_id, branch_id, cycle_id)

    # Buscar avaliações detalhadas
    emp_q = db.table("performance_employees").select("id,name,cargo,hierarchy_level,performance_branches(name),performance_companies(name)").eq("active", True)
    if company_id:
        emp_q = emp_q.eq("company_id", company_id)
    if branch_id:
        emp_q = emp_q.eq("branch_id", branch_id)
    employees = emp_q.execute().data
    emp_map = {e["id"]: e for e in employees}

    reviews = (
        db.table("performance_reviews")
        .select("id,employee_id,evaluator_id,final_score,status,is_self_evaluation,submitted_at,observations")
        .eq("cycle_id", cycle_id)
        .execute()
        .data
    )

    self_eval_map = {r["employee_id"]: r for r in reviews if r.get("is_self_evaluation")}
    manager_reviews = [r for r in reviews if not r.get("is_self_evaluation")]

    # Calibrações
    review_ids = [r["id"] for r in manager_reviews]
    calib_map: dict[str, dict] = {}
    if review_ids:
        calibs = db.table("performance_calibrations").select("review_id,calibrated_by,calibrated_at,notes").in_("review_id", review_ids).order("calibrated_at", desc=True).execute().data
        for c in calibs:
            if c["review_id"] not in calib_map:
                calib_map[c["review_id"]] = c

    wb = Workbook()

    # ── Paleta de cores ─────────────────────────────────────────────────────────
    GREEN  = "00694E"
    LGREEN = "E6F4F0"
    VIOLET = "7C3AED"
    AMBER  = "D97706"
    GRAY   = "F3F4F6"
    WHITE  = "FFFFFF"

    def header_style(cell, bg=GREEN, bold=True, color=WHITE):
        cell.font = Font(bold=bold, color=color, size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def thin_border():
        s = Side(style="thin", color="DDDDDD")
        return Border(left=s, right=s, top=s, bottom=s)

    # ── Aba 1: Resumo ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16

    cycle_name = "—"
    if cycle_id:
        cy = db.table("performance_cycles").select("name").eq("id", cycle_id).execute()
        if cy.data:
            cycle_name = cy.data[0]["name"]

    rows = [
        ("Ciclo", cycle_name),
        ("Total Avaliados", dash.get("total_evaluated", 0)),
        ("Completude (%)", f"{dash.get('completion_pct', 0)}%"),
        ("Pendentes Ciência", dash.get("pending_acknowledgment", 0)),
        ("Sem Avaliação", dash.get("without_evaluation", 0)),
        ("Auto-Avaliações Enviadas", dash.get("self_eval_sent", 0)),
        ("Auto-Avaliações Concluídas", dash.get("self_eval_completed", 0)),
        ("Auto-Avaliações (%)", f"{dash.get('self_eval_pct', 0)}%"),
        ("Calibrações Realizadas", dash.get("calibrations_count", 0)),
        ("Calibrações (%)", f"{dash.get('calibrations_pct', 0)}%"),
    ]
    ws.append(["Indicador", "Valor"])
    for cell in ws[1]:
        header_style(cell)
    for r in rows:
        ws.append(list(r))
        for cell in ws[ws.max_row]:
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Aba 2: Avaliações ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Avaliações")
    headers2 = ["Colaborador", "Cargo", "Nível", "Empresa", "Filial", "Avaliador", "Nota Final", "Status", "Calibrado por", "Obs. Gestor"]
    ws2.append(headers2)
    for cell in ws2[1]:
        header_style(cell)

    _PERFIL_LABEL_XLS = {"gerente": "Gerente", "coordenador_supervisor": "Coord./Supervisor",
                         "administrativo": "Administrativo", "operacional": "Operacional",
                         }
    _HLEVEL_LABEL_XLS = {1: "Gerente", 2: "Coord./Supervisor", 3: "Adm./Operacional"}
    def _emp_level_label(emp_data: dict) -> str:
        p = emp_data.get("perfil") or ""
        return _PERFIL_LABEL_XLS.get(p) or _HLEVEL_LABEL_XLS.get(emp_data.get("hierarchy_level"), "")

    status_label = {"pending": "Pendente", "completed": "Avaliado", "acknowledged": "Ciência Dada", "calibrated": "Calibrado"}

    for r in manager_reviews:
        emp = emp_map.get(r.get("employee_id", ""), {})
        ev = emp_map.get(r.get("evaluator_id", ""), {})
        branch = emp.get("performance_branches") or {}
        company = emp.get("performance_companies") or {}
        calib = calib_map.get(r["id"])
        ws2.append([
            emp.get("name", ""),
            emp.get("cargo", ""),
            _emp_level_label(emp),
            company.get("name", "") if isinstance(company, dict) else "",
            branch.get("name", "") if isinstance(branch, dict) else "",
            ev.get("name", ""),
            float(r["final_score"]) if r.get("final_score") is not None else "",
            status_label.get(r.get("status", ""), r.get("status", "")),
            calib["calibrated_by"] if calib else "",
            r.get("observations") or "",
        ])
        for cell in ws2[ws2.max_row]:
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for i, col in enumerate(headers2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = max(14, len(col) + 4)

    # ── Aba 3: Auto-Avaliações ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Auto-Avaliações")
    headers3 = ["Colaborador", "Cargo", "Nível", "Empresa", "Filial", "Status Auto-Aval.", "Obs. Colaborador"]
    ws3.append(headers3)
    for cell in ws3[1]:
        header_style(cell, bg=VIOLET)

    for emp in sorted(employees, key=lambda e: e.get("name", "")):
        se = self_eval_map.get(emp["id"])
        branch = emp.get("performance_branches") or {}
        company = emp.get("performance_companies") or {}
        ws3.append([
            emp.get("name", ""),
            emp.get("cargo", ""),
            _emp_level_label(emp),
            company.get("name", "") if isinstance(company, dict) else "",
            branch.get("name", "") if isinstance(branch, dict) else "",
            "Concluída" if se else "Pendente",
            se.get("observations") or "" if se else "",
        ])
        row_cells = ws3[ws3.max_row]
        fill_color = LGREEN if se else "FEF3C7"
        for cell in row_cells:
            cell.border = thin_border()
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for i, col in enumerate(headers3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = max(14, len(col) + 4)

    # ── Aba 4: Médias por Indicador ──────────────────────────────────────────────
    ws4 = wb.create_sheet("Médias por Indicador")
    ws4.append(["Indicador", "Média"])
    for cell in ws4[1]:
        header_style(cell, bg=AMBER)
    for ind in dash.get("indicator_averages", []):
        ws4.append([ind["name"], ind["avg"]])
        for cell in ws4[ws4.max_row]:
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws4.column_dimensions["A"].width = 40
    ws4.column_dimensions["B"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"desempenho_{cycle_name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/dashboard/pending-ciencia")
def dashboard_pending_ciencia(
    _: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
    cycle_id: str | None = None,
    company_id: str | None = None,
) -> list[dict]:
    """Employees who have a completed review but haven't acknowledged yet."""
    db = get_supabase()
    if not cycle_id:
        cycle = _get_current_cycle(db)
        cycle_id = cycle["id"] if cycle else None
    if not cycle_id:
        return []

    rev_q = (
        db.table("performance_reviews")
        .select("id,employee_id,evaluator_id,final_score")
        .eq("cycle_id", cycle_id)
        .in_("status", ["completed", "calibrated"])
    )
    reviews = rev_q.execute().data
    if not reviews:
        return []

    # Filter by company if needed
    if company_id:
        emp_res = db.table("performance_employees").select("id").eq("company_id", company_id).execute().data
        company_emp_ids = {e["id"] for e in emp_res}
        reviews = [r for r in reviews if r.get("employee_id") in company_emp_ids]
        if not reviews:
            return []

    review_ids = [r["id"] for r in reviews]
    acks = (
        db.table("performance_review_acknowledgments")
        .select("review_id")
        .in_("review_id", review_ids)
        .execute()
        .data
    )
    acked_review_ids = {a["review_id"] for a in acks}

    pending_reviews = [r for r in reviews if r["id"] not in acked_review_ids]
    if not pending_reviews:
        return []

    # Fetch employee and evaluator names
    all_emp_ids = list({r["employee_id"] for r in pending_reviews} | {r["evaluator_id"] for r in pending_reviews if r.get("evaluator_id")})
    people = db.table("performance_employees").select("id,name,cargo").in_("id", all_emp_ids).execute().data
    people_map = {p["id"]: p for p in people}

    result = []
    for rev in pending_reviews:
        emp = people_map.get(rev["employee_id"], {})
        evaluator = people_map.get(rev.get("evaluator_id", ""), {})
        result.append({
            "employee_name": emp.get("name", ""),
            "employee_cargo": emp.get("cargo", ""),
            "evaluator_name": evaluator.get("name", ""),
            "final_score": rev.get("final_score"),
        })
    result.sort(key=lambda x: x["employee_name"])
    return result


# ── Cycles (list) ─────────────────────────────────────────────────────────────

@router.get("/cycles")
def list_cycles(_: Annotated[dict, Depends(require_role(*_RH_ADMIN))]) -> list[dict]:
    """Retorna todos os ciclos (usado pelo filtro de ciclo no dashboard)."""
    cached = _cache_get("cycles")
    if cached is not None:
        return cached
    db = get_supabase()
    result = (
        db.table("performance_cycles")
        .select("id,name,status,period_start,period_end,created_at")
        .order("created_at", desc=True)
        .execute()
        .data
    )
    _cache_set("cycles", result)
    return result


# ── Companies / Branches / Employees ──────────────────────────────────────────

@router.get("/companies")
def list_companies(_: Annotated[dict, Depends(require_role(*_RH_ADMIN))]) -> list[dict]:
    cached = _cache_get("companies")
    if cached is not None:
        return cached
    db = get_supabase()
    result = db.table("performance_companies").select("*").eq("active", True).order("name").execute().data
    _cache_set("companies", result)
    return result


@router.get("/branches")
def list_branches(
    _: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
    company_id: str | None = None,
) -> list[dict]:
    cache_key = f"branches:{company_id or 'all'}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached
    db = get_supabase()
    q = db.table("performance_branches").select("*").eq("active", True)
    if company_id:
        q = q.eq("company_id", company_id)
    result = q.order("name").execute().data
    _cache_set(cache_key, result)
    return result


@router.get("/employees/template")
def download_template(_: Annotated[dict, Depends(require_role(*_RH_ADMIN))]):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        raise HTTPException(500, detail="openpyxl não instalado")

    db = get_supabase()
    companies = db.table("performance_companies").select("id,name").eq("active", True).order("name").execute().data
    branches_all = db.table("performance_branches").select("id,name,company_id").eq("active", True).order("name").execute().data

    branches_by_company: dict[str, list[str]] = {}
    for b in branches_all:
        branches_by_company.setdefault(b["company_id"], []).append(b["name"])

    company_names = [c["name"] for c in companies]
    all_branch_names = sorted({b["name"] for b in branches_all})

    wb = openpyxl.Workbook()
    ws_lists = wb.active
    ws_lists.title = "Listas"

    # Column A: company names
    ws_lists["A1"] = "Empresa"
    ws_lists["A1"].font = Font(bold=True)
    for i, cn in enumerate(company_names, 2):
        ws_lists.cell(row=i, column=1).value = cn

    # Columns B onwards: branches per company (one column each)
    for col_i, company in enumerate(companies, 2):
        ws_lists.cell(row=1, column=col_i).value = f"Filiais: {company['name']}"
        ws_lists.cell(row=1, column=col_i).font = Font(bold=True)
        for row_i, b_name in enumerate(branches_by_company.get(company["id"], []), 2):
            ws_lists.cell(row=row_i, column=col_i).value = b_name

    # Column for all branches combined (used for validation)
    all_col = len(companies) + 2
    ws_lists.cell(row=1, column=all_col).value = "Todas as Filiais"
    ws_lists.cell(row=1, column=all_col).font = Font(bold=True)
    for i, b_name in enumerate(all_branch_names, 2):
        ws_lists.cell(row=i, column=all_col).value = b_name

    # Level and yes/no columns
    lv_col = all_col + 1
    yn_col = all_col + 2
    ws_lists.cell(row=1, column=lv_col).value = "Nível Hierárquico"
    ws_lists.cell(row=1, column=lv_col).font = Font(bold=True)
    for i, lv in enumerate(["Gerente", "Coordenador-Supervisor", "Administrativo", "Operacional"], 2):
        ws_lists.cell(row=i, column=lv_col).value = lv
    ws_lists.cell(row=1, column=yn_col).value = "Tem E-mail"
    ws_lists.cell(row=1, column=yn_col).font = Font(bold=True)
    ws_lists.cell(row=2, column=yn_col).value = "Sim"
    ws_lists.cell(row=3, column=yn_col).value = "Não"

    # Named range reference for all branches column
    all_branches_col_letter = get_column_letter(all_col)
    all_branches_ref = f"Listas!${all_branches_col_letter}$2:${all_branches_col_letter}${len(all_branch_names) + 1}"
    company_list_ref = f"Listas!$A$2:$A${len(company_names) + 1}"
    level_ref = f"Listas!${get_column_letter(lv_col)}$2:${get_column_letter(lv_col)}$5"
    yn_ref = f"Listas!${get_column_letter(yn_col)}$2:${get_column_letter(yn_col)}$3"

    # ── Instructions sheet ────────────────────────────────────────────────────
    ws_inst = wb.create_sheet("Instruções")
    ws_inst["A1"] = "COMO PREENCHER O TEMPLATE"
    ws_inst["A1"].font = Font(bold=True, size=14)
    ws_inst["A3"] = "Estrutura das colunas (9 colunas — A a I):"
    ws_inst["A3"].font = Font(bold=True, size=12)
    col_desc = [
        ("A", "Empresa",                  "Obrigatório — selecione da lista suspensa."),
        ("B", "Filial",                   "Obrigatório — selecione da lista suspensa."),
        ("C", "Gerência",                 "Obrigatório — nome da área. Ex: 'Gerência de Recursos Humanos'"),
        ("D", "Nome Completo",            "Obrigatório — sem abreviações. Mín. 2 palavras com 3+ letras."),
        ("E", "Cargo",                    "Obrigatório — cargo do colaborador. Ex: 'Analista de RH Pleno'"),
        ("F", "Nível Hierárquico",        "Obrigatório — selecione da lista: Gerente / Coordenador-Supervisor / Administrativo / Operacional"),
        ("G", "E-mail Corporativo",       "Obrigatório se Tem E-mail = Sim; deixe em branco se Não."),
        ("H", "Tem E-mail Corporativo?",  "Obrigatório — Sim ou Não."),
        ("I", "CPF",                      "Obrigatório se Tem E-mail = Não. 11 dígitos numéricos sem pontos/traços. Ex: 12345678901"),
    ]
    for row_i, (col, header, desc) in enumerate(col_desc, 4):
        ws_inst.cell(row=row_i, column=1).value = col
        ws_inst.cell(row=row_i, column=1).font = Font(bold=True)
        ws_inst.cell(row=row_i, column=2).value = header
        ws_inst.cell(row=row_i, column=2).font = Font(bold=True)
        ws_inst.cell(row=row_i, column=3).value = desc

    ws_inst["A14"] = "REGRAS IMPORTANTES"
    ws_inst["A14"].font = Font(bold=True, size=12)
    rules = [
        "- CPF é obrigatório para todos os colaboradores (avaliadores e avaliados).",
        "- Colaboradores SEM e-mail corporativo fazem ciência presencial usando o CPF.",
        "- Gerentes (nível 1) são avaliadores; Coordenadores/Supervisores (nível 2) avaliam e são avaliados.",
        "- Administrativo e Operacional são perfis distintos (nível 3), cada um com indicadores de avaliação próprios.",
        "- A hierarquia (quem avalia quem) é configurada manualmente no sistema após a importação.",
        "- Filiais: os nomes devem ser exatamente como aparecem na aba Listas (lista suspensa).",
    ]
    for i, rule in enumerate(rules, 15):
        ws_inst.cell(row=i, column=1).value = rule
    ws_inst.column_dimensions["A"].width = 8
    ws_inst.column_dimensions["B"].width = 30
    ws_inst.column_dimensions["C"].width = 80

    ws_inst["A22"] = "EXEMPLO:"
    ws_inst["A22"].font = Font(bold=True, size=12)
    ex_headers = ["A:Empresa", "B:Filial", "C:Gerência", "D:Nome", "E:Cargo", "F:Nível", "G:E-mail", "H:Tem E-mail?", "I:CPF"]
    for ci, h in enumerate(ex_headers, 1):
        ws_inst.cell(row=23, column=ci).value = h
        ws_inst.cell(row=23, column=ci).font = Font(bold=True)
    examples = [
        [company_names[0] if company_names else "Empresa", all_branch_names[0] if all_branch_names else "Filial",
         "Gerência de Operações", "João da Silva Santos",
         "Gerente de Operações", "Gerente", "joao.santos@empresa.com.br", "Sim", "07123456789"],
        [company_names[0] if company_names else "Empresa", all_branch_names[0] if all_branch_names else "Filial",
         "Gerência de Operações", "Maria Fernanda Costa",
         "Coordenadora de Logística", "Coordenador-Supervisor", "maria.costa@empresa.com.br", "Sim", "04987654321"],
        [company_names[0] if company_names else "Empresa", all_branch_names[0] if all_branch_names else "Filial",
         "Gerência de Operações", "Carlos Eduardo Lima",
         "Operador Logístico", "Operacional", "", "Não", "12345678901"],
    ]
    for r_i, row_data in enumerate(examples, 24):
        for c_i, val in enumerate(row_data, 1):
            ws_inst.cell(row=r_i, column=c_i).value = val

    # ── Dados sheet ──────────────────────────────────────────────────────────
    ws_data = wb.create_sheet("Dados")
    headers = [
        "Empresa", "Filial", "Gerência", "Nome Completo", "Cargo",
        "Nível Hierárquico", "E-mail Corporativo", "Tem E-mail Corporativo?", "CPF",
    ]
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF")
    for i, h in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=i)
        cell.value = h
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    instrucoes = [
        "Obrigatório — selecione da lista",
        "Obrigatório — selecione da lista",
        "Obrigatório — ex: Gerência de Operações",
        "Obrigatório — nome completo sem abreviações",
        "Obrigatório — cargo do colaborador",
        "Obrigatório — selecione da lista",
        "Obrigatório se Tem E-mail = Sim; em branco se Não",
        "Obrigatório — Sim ou Não",
        "Obrigatório se Tem E-mail = Não. 11 dígitos sem pontos/traços. Ex: 12345678901",
    ]
    inst_font = Font(italic=True, color="808080")
    for i, inst in enumerate(instrucoes, 1):
        ws_data.cell(row=2, column=i).value = inst
        ws_data.cell(row=2, column=i).font = inst_font

    # Data validations
    dv_empresa = DataValidation(type="list", formula1=f"={company_list_ref}", allow_blank=False, showErrorMessage=True,
                                errorTitle="Empresa inválida", error="Selecione uma empresa da lista.")
    dv_filial = DataValidation(type="list", formula1=f"={all_branches_ref}", allow_blank=False, showErrorMessage=True,
                               errorTitle="Filial inválida", error="Selecione uma filial da lista. O nome deve ser idêntico ao sistema.")
    dv_nivel = DataValidation(type="list", formula1=f"={level_ref}", allow_blank=False, showErrorMessage=True,
                              errorTitle="Nível inválido", error="Use: Gerente, Coordenador-Supervisor, Administrativo ou Operacional")
    dv_email = DataValidation(type="list", formula1=f"={yn_ref}", allow_blank=False)
    ws_data.add_data_validation(dv_empresa)
    ws_data.add_data_validation(dv_filial)
    ws_data.add_data_validation(dv_nivel)
    ws_data.add_data_validation(dv_email)
    dv_empresa.sqref = "A3:A502"
    dv_filial.sqref = "B3:B502"
    dv_nivel.sqref = "F3:F502"
    dv_email.sqref = "H3:H502"

    col_widths = [30, 28, 28, 40, 35, 28, 38, 26, 18]
    for i, w in enumerate(col_widths, 1):
        ws_data.column_dimensions[get_column_letter(i)].width = w
    ws_data.row_dimensions[1].height = 40
    ws_data.row_dimensions[2].height = 25

    wb.move_sheet("Dados", offset=-len(wb.worksheets))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_hierarquia_avd.xlsx"},
    )


@router.get("/employees")
def list_employees(
    _: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
    company_id: str | None = None,
    branch_id: str | None = None,
) -> list[dict]:
    db = get_supabase()
    q = db.table("performance_employees").select("*").eq("active", True)
    if company_id:
        q = q.eq("company_id", company_id)
    if branch_id:
        q = q.eq("branch_id", branch_id)
    employees = q.order("name").execute().data

    # Busca apenas os gestores necessários (evita varredura completa da tabela)
    manager_ids = list({e["manager_id"] for e in employees if e.get("manager_id")})
    mgr_map: dict[str, str] = {}
    if manager_ids:
        mgrs = (
            db.table("performance_employees")
            .select("id,name")
            .in_("id", manager_ids)
            .execute()
            .data
        )
        mgr_map = {m["id"]: m["name"] for m in mgrs}

    result = []
    for e in employees:
        out = _emp_out(e)
        out["manager_name"] = mgr_map.get(e.get("manager_id"), "") if e.get("manager_id") else ""
        result.append(out)
    return result


class EmployeeBody(BaseModel):
    name: str
    matricula: str = ""          # mantido na DB mas não exibido/exigido na UI
    cargo: str = ""
    level: str = "administrativo_operacional"
    email: str | None = None
    cpf: str | None = None       # obrigatório quando não há e-mail
    has_corporate_email: bool = True
    whatsapp_phone: str = ""
    manager_id: str | None = None
    branch_id: str | None = None
    company_id: str | None = None
    management_id: str | None = None
    jarvis_username: str | None = None
    active: bool | None = None

    @property
    def perfil(self) -> str:
        return self.level if self.level in _PERFIS_VALIDOS else ""


_PREPS = {"da", "de", "di", "do", "dos", "das", "e"}

def _validate_name(name: str) -> None:
    parts = [p for p in name.strip().split() if p]
    meaningful = [p for p in parts if p.lower() not in _PREPS]
    if len(parts) < 2 or len(meaningful) < 2:
        raise HTTPException(400, detail="Nome deve ser completo, sem abreviações (mínimo 2 partes com 3+ caracteres)")


def _validate_cpf(cpf: str) -> str:
    """Valida CPF brasileiro (dígitos verificadores mod-11). Retorna CPF limpo (só dígitos)."""
    digits = re.sub(r'\D', '', cpf)
    if len(digits) != 11:
        raise HTTPException(400, detail="CPF deve ter 11 dígitos numéricos")
    if len(set(digits)) == 1:   # ex: 00000000000, 11111111111 ...
        raise HTTPException(400, detail="CPF inválido")
    # 1º dígito verificador
    total = sum(int(digits[i]) * (10 - i) for i in range(9))
    d1 = 0 if total % 11 < 2 else 11 - (total % 11)
    if int(digits[9]) != d1:
        raise HTTPException(400, detail="CPF inválido — dígito verificador incorreto")
    # 2º dígito verificador
    total = sum(int(digits[i]) * (11 - i) for i in range(10))
    d2 = 0 if total % 11 < 2 else 11 - (total % 11)
    if int(digits[10]) != d2:
        raise HTTPException(400, detail="CPF inválido — dígito verificador incorreto")
    return digits


def _ensure_management(db, branch_id: str | None) -> str | None:
    if not branch_id:
        return None
    mgmt = (
        db.table("performance_managements")
        .select("id")
        .eq("branch_id", branch_id)
        .eq("active", True)
        .limit(1)
        .execute()
    )
    if mgmt.data:
        return mgmt.data[0]["id"]
    res = db.table("performance_managements").insert(
        {"branch_id": branch_id, "name": "Geral", "active": True}
    ).execute()
    return res.data[0]["id"] if res.data else None


@router.post("/employees", status_code=status.HTTP_201_CREATED)
def create_employee(
    body: EmployeeBody,
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    _validate_name(body.name)
    has_email = bool(body.email and body.email.strip())

    # CPF obrigatório para todos os colaboradores; validado com dígito verificador
    if not body.cpf or not body.cpf.strip():
        raise HTTPException(400, detail="CPF é obrigatório para todos os colaboradores")
    cpf_clean = _validate_cpf(body.cpf)

    # Verificar CPF duplicado
    if cpf_clean:
        dup_cpf = db_dup = get_supabase() if False else None  # lazy init
        db = get_supabase()
        dup = db.table("performance_employees").select("id,name").eq("cpf", cpf_clean).eq("active", True).execute()
        if dup.data:
            raise HTTPException(400, detail=f"CPF já cadastrado para: {dup.data[0]['name']}")
    else:
        db = get_supabase()

    hierarchy_level = _LEVEL_MAP.get(body.level, 3)
    mgmt_id = body.management_id or _ensure_management(db, body.branch_id)

    # Matrícula: usar CPF quando disponível, senão gerar sequencial único
    import uuid as _uuid
    matricula = cpf_clean or _uuid.uuid4().hex[:10].upper()

    result = db.table("performance_employees").insert({
        "name": body.name.strip(),
        "matricula": matricula,
        "email": body.email or None,
        "cpf": cpf_clean,
        "cargo": body.cargo or "",
        "has_corporate_email": has_email,
        "whatsapp_phone": body.whatsapp_phone or "",
        "perfil": body.perfil,
        "hierarchy_level": hierarchy_level,
        "manager_id": body.manager_id or None,
        "management_id": mgmt_id,
        "branch_id": body.branch_id,
        "company_id": body.company_id,
        "jarvis_username": body.jarvis_username,
        "active": True,
    }).execute()
    if not result.data:
        raise HTTPException(500, detail="Erro ao cadastrar colaborador")
    emp = result.data[0]
    log_action("employee", emp["id"], "create", None, emp, current_user["username"], request)
    return _emp_out(emp)


@router.put("/employees/{employee_id}")
def update_employee(
    employee_id: str,
    body: EmployeeBody,
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    db = get_supabase()
    existing = db.table("performance_employees").select("*").eq("id", employee_id).execute()
    if not existing.data:
        raise HTTPException(404, detail="Colaborador não encontrado")
    old = existing.data[0]

    updates: dict[str, Any] = {"updated_at": "now()"}
    if body.name:
        _validate_name(body.name)
        updates["name"] = body.name.strip()
    if body.matricula:
        # matricula agora é gerada automaticamente a partir do CPF — aceita sem validação extra
        updates["matricula"] = body.matricula.strip()
    if body.cargo:
        updates["cargo"] = body.cargo
    if body.email is not None:
        updates["email"] = body.email or None
        updates["has_corporate_email"] = bool(body.email and body.email.strip())
    if body.cpf is not None:
        updates["cpf"] = re.sub(r'\D', '', body.cpf).strip() or None
    if body.whatsapp_phone is not None:
        updates["whatsapp_phone"] = re.sub(r'\D', '', body.whatsapp_phone or "").strip()
    if body.level:
        updates["perfil"] = body.perfil
        updates["hierarchy_level"] = _LEVEL_MAP.get(body.level, 3)
    if body.manager_id is not None:
        updates["manager_id"] = body.manager_id or None
    if body.company_id:
        updates["company_id"] = body.company_id
    if body.branch_id:
        updates["branch_id"] = body.branch_id
    if body.active is not None:
        updates["active"] = body.active
    if body.jarvis_username is not None:
        updates["jarvis_username"] = body.jarvis_username

    result = db.table("performance_employees").update(updates).eq("id", employee_id).execute()
    if not result.data:
        raise HTTPException(500, detail="Erro ao atualizar colaborador")
    new = result.data[0]
    log_action("employee", employee_id, "update", old, new, current_user["username"], request)
    return _emp_out(new)


@router.post("/employees/import")
async def import_employees(
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
    file: UploadFile = File(...),
) -> dict:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, detail="openpyxl não instalado")

    if not (file.filename or "").endswith(".xlsx"):
        raise HTTPException(400, detail="Arquivo deve ser .xlsx")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception:
        raise HTTPException(400, detail="Arquivo inválido ou corrompido")

    if "Dados" not in wb.sheetnames:
        raise HTTPException(400, detail="Aba 'Dados' não encontrada. Use o template correto.")

    ws = wb["Dados"]
    db = get_supabase()

    companies = {c["name"]: c for c in db.table("performance_companies").select("*").execute().data}
    branches_all = db.table("performance_branches").select("*").execute().data
    existing_cpfs = {
        e["cpf"]: True
        for e in db.table("performance_employees").select("cpf").execute().data
        if e.get("cpf")
    }
    NIVEL_MAP = {
        "gerente": 1,
        "coordenador-supervisor": 2,
        "administrativo": 3,
        "operacional": 3,
    }
    NIVEL_PERFIL = {
        "gerente": "gerente",
        "coordenador-supervisor": "coordenador_supervisor",
        "administrativo": "administrativo",
        "operacional": "operacional",
    }

    errors: list[dict] = []
    rows_data: list[dict] = []
    file_cpfs: dict[str, int] = {}   # cpf → first line seen (dedup within file)

    # Detecta se dados começam na linha 2 (1 header) ou linha 3 (header + instrução)
    company_names_set = set(companies.keys())
    row2_val = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
    row2_empresa = str(row2_val[0] if row2_val else "").strip()
    start_row = 2 if row2_empresa in company_names_set else 3

    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if all(v is None or str(v).strip() == "" for v in row):
            break

        # ── Novo formato de 9 colunas (sem Matrícula, sem Gestor) ──────────
        empresa_name  = str(row[0] or "").strip()          # A
        filial_name   = str(row[1] or "").strip()          # B
        gerencia_name = str(row[2] or "").strip()          # C
        nome          = str(row[3] or "").strip()          # D
        cargo         = str(row[4] or "").strip()          # E
        nivel_str     = str(row[5] or "").strip()          # F
        email         = str(row[6] or "").strip() or None  # G
        tem_email_str = str(row[7] or "").strip()          # H
        cpf_raw       = re.sub(r'\D', '', str(row[8] or "")) if len(row) > 8 else ""  # I
        cpf           = cpf_raw if cpf_raw else None

        row_errors: list[dict] = []

        # Empresa
        company = companies.get(empresa_name)
        if not company:
            row_errors.append({"linha": row_idx, "campo": "Empresa", "erro": f"Empresa '{empresa_name}' inválida"})

        # Filial
        branch = None
        if company:
            branch_matches = [b for b in branches_all if b["company_id"] == company["id"] and b["name"].lower() == filial_name.lower()]
            if not branch_matches:
                row_errors.append({"linha": row_idx, "campo": "Filial", "erro": f"Filial '{filial_name}' não encontrada. Nome deve ser idêntico ao sistema."})
            else:
                branch = branch_matches[0]

        # Nome — permite preposições curtas brasileiras (da, de, do, dos, das)
        _PREPS = {"da", "de", "di", "do", "dos", "das", "e"}
        parts = [p for p in nome.split() if p]
        meaningful = [p for p in parts if p.lower() not in _PREPS]
        if len(parts) < 2 or len(meaningful) < 2:
            row_errors.append({"linha": row_idx, "campo": "Nome Completo", "erro": "Nome deve ser completo, sem abreviações (mín. 2 palavras com 3+ letras)"})

        # Nível
        nivel_key = nivel_str.lower()
        hierarchy_level = NIVEL_MAP.get(nivel_key)
        nivel_perfil = NIVEL_PERFIL.get(nivel_key, "administrativo_operacional")
        if hierarchy_level is None:
            row_errors.append({"linha": row_idx, "campo": "Nível Hierárquico", "erro": f"Nível '{nivel_str}' inválido. Use: Gerente / Coordenador-Supervisor / Administrativo / Operacional"})

        # E-mail
        tem_email = tem_email_str.lower() in ("sim", "s", "yes")
        if tem_email and not email:
            row_errors.append({"linha": row_idx, "campo": "E-mail Corporativo", "erro": "E-mail obrigatório quando Tem E-mail = Sim"})

        # CPF — obrigatório para todos
        if not cpf:
            row_errors.append({"linha": row_idx, "campo": "CPF", "erro": "CPF é obrigatório para todos os colaboradores. 11 dígitos numéricos."})
        elif len(cpf) != 11:
            row_errors.append({"linha": row_idx, "campo": "CPF", "erro": f"CPF deve ter 11 dígitos numéricos (encontrado: {len(cpf)})"})
        else:
            if cpf in existing_cpfs:
                row_errors.append({"linha": row_idx, "campo": "CPF", "erro": f"CPF {cpf} já cadastrado no sistema"})
            elif cpf in file_cpfs:
                row_errors.append({"linha": row_idx, "campo": "CPF", "erro": f"CPF {cpf} duplicado no arquivo (linha {file_cpfs[cpf]})"})
            else:
                file_cpfs[cpf] = row_idx

        if row_errors:
            errors.extend(row_errors)
        else:
            # Matrícula auto-gerada a partir do CPF
            import uuid as _uuid
            matricula = cpf if cpf else f"IMP{str(_uuid.uuid4().int)[:8]}"
            rows_data.append({
                "company_id": company["id"] if company else None,
                "branch_id": branch["id"] if branch else None,
                "gerencia_name": gerencia_name,
                "name": nome, "matricula": matricula, "cargo": cargo,
                "hierarchy_level": hierarchy_level,
                "perfil": nivel_perfil,
                "email": email,
                "has_corporate_email": tem_email,
                "cpf": cpf,
            })

    if errors:
        return {"errors": [f"Linha {e['linha']} — {e['campo']}: {e['erro']}" for e in errors], "imported": 0}
    if not rows_data:
        return {"errors": ["Nenhuma linha de dados encontrada (preencha a partir da linha 3)"], "imported": 0}

    existing_mgmts = {
        f"{m['branch_id']}_{m['name'].lower()}": m["id"]
        for m in db.table("performance_managements").select("*").execute().data
    }
    imported = 0

    for row in sorted(rows_data, key=lambda x: x["hierarchy_level"]):
        mgmt_key = f"{row['branch_id']}_{row['gerencia_name'].lower()}"
        mgmt_id = existing_mgmts.get(mgmt_key)
        if not mgmt_id:
            mgmt_res = db.table("performance_managements").insert(
                {"branch_id": row["branch_id"], "name": row["gerencia_name"], "active": True}
            ).execute()
            if mgmt_res.data:
                mgmt_id = mgmt_res.data[0]["id"]
                existing_mgmts[mgmt_key] = mgmt_id

        emp_res = db.table("performance_employees").insert({
            "name": row["name"], "matricula": row["matricula"],
            "email": row["email"], "cargo": row["cargo"],
            "has_corporate_email": row["has_corporate_email"],
            "cpf": row.get("cpf"),
            "hierarchy_level": row["hierarchy_level"],
            "perfil": row.get("perfil", "administrativo_operacional"),
            "manager_id": None,           # hierarquia definida via UI
            "management_id": mgmt_id,
            "branch_id": row["branch_id"], "company_id": row["company_id"],
            "active": True,
        }).execute()
        if emp_res.data:
            imported += 1

    log_action("system", "import-excel", "bulk_import", None, {"imported": imported}, current_user["username"], request)
    return {"errors": [], "imported": imported}


# ── Current cycle management ───────────────────────────────────────────────────

@router.get("/cycle/status")
def get_cycle_status(_: Annotated[dict, Depends(require_role(*_RH_ADMIN))]) -> dict | None:
    db = get_supabase()
    cycle = _get_current_cycle(db)
    if not cycle:
        return None
    company_name = None
    if cycle.get("company_id"):
        co = db.table("performance_companies").select("name").eq("id", cycle["company_id"]).execute()
        company_name = co.data[0]["name"] if co.data else None
    return {
        "id": cycle["id"],
        "name": cycle["name"],
        "status": cycle["status"],
        "is_open": cycle["status"] == "open",
        "started_at": cycle.get("created_at"),
        "period_start": cycle.get("period_start"),
        "period_end": cycle.get("period_end"),
        "company_id": cycle.get("company_id"),
        "company_name": company_name,
    }


class CycleCreateSimple(BaseModel):
    name: str
    period_start: str | None = None  # ISO date string
    period_end: str | None = None
    company_id: str | None = None  # None = todas as empresas


@router.post("/cycle", status_code=status.HTTP_201_CREATED)
def create_cycle(
    body: CycleCreateSimple,
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    if not body.name.strip():
        raise HTTPException(400, detail="Nome do ciclo é obrigatório")
    db = get_supabase()
    open_cycle = db.table("performance_cycles").select("id,name").eq("status", "open").execute()
    if open_cycle.data:
        raise HTTPException(400, detail=f"Já existe um ciclo aberto: '{open_cycle.data[0]['name']}'")
    today = date.today()
    insert_data: dict = {
        "name": body.name.strip(),
        "period_start": body.period_start or today.isoformat(),
        "period_end": body.period_end or (today + timedelta(days=90)).isoformat(),
        "status": "draft",
        "created_by": current_user["username"],
    }
    if body.company_id:
        insert_data["company_id"] = body.company_id
    result = db.table("performance_cycles").insert(insert_data).execute()
    if not result.data:
        raise HTTPException(500, detail="Erro ao criar ciclo")
    cycle = result.data[0]
    _cache_invalidate_prefix("cycles")
    _cache_invalidate_prefix("dashboard:")
    log_action("cycle", cycle["id"], "create", None, cycle, current_user["username"], request)
    company_name = None
    if body.company_id:
        co = db.table("performance_companies").select("name").eq("id", body.company_id).execute()
        company_name = co.data[0]["name"] if co.data else None
    return {
        "id": cycle["id"], "name": cycle["name"],
        "status": cycle["status"], "is_open": False,
        "started_at": cycle.get("created_at"),
        "period_start": cycle.get("period_start"),
        "period_end": cycle.get("period_end"),
        "company_id": body.company_id,
        "company_name": company_name,
    }


@router.post("/cycle/open")
def open_current_cycle(
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    db = get_supabase()
    cycle = _get_current_cycle(db)
    if not cycle:
        raise HTTPException(404, detail="Nenhum ciclo encontrado")
    if cycle["status"] != "draft":
        raise HTTPException(400, detail="Apenas ciclos em rascunho podem ser abertos")
    db.table("performance_cycles").update({"status": "open"}).eq("id", cycle["id"]).execute()
    _cache_invalidate_prefix("cycles")
    _cache_invalidate_prefix("dashboard:")
    log_action("cycle", cycle["id"], "open", {"status": "draft"}, {"status": "open"}, current_user["username"], request)
    return {"ok": True, "status": "open", "is_open": True}


@router.post("/cycle/close")
def close_current_cycle(
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    db = get_supabase()
    cycle = _get_current_cycle(db)
    if not cycle:
        raise HTTPException(404, detail="Nenhum ciclo encontrado")
    if cycle["status"] != "open":
        raise HTTPException(400, detail="Apenas ciclos abertos podem ser encerrados")
    db.table("performance_evaluation_tokens").update({
        "invalidated_at": datetime.now(tz=timezone.utc).isoformat(),
    }).eq("cycle_id", cycle["id"]).eq("is_used", False).is_("invalidated_at", "null").execute()
    db.table("performance_cycles").update({"status": "closed"}).eq("id", cycle["id"]).execute()
    _cache_invalidate_prefix("cycles")
    _cache_invalidate_prefix("dashboard:")
    log_action("cycle", cycle["id"], "close", {"status": "open"}, {"status": "closed"}, current_user["username"], request)
    return {"ok": True, "status": "closed", "is_open": False}


class ReopenBody(BaseModel):
    justification: str


@router.post("/cycle/reopen")
def reopen_current_cycle(
    body: ReopenBody,
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    if not body.justification.strip():
        raise HTTPException(400, detail="Justificativa é obrigatória")
    db = get_supabase()
    closed = (
        db.table("performance_cycles")
        .select("*")
        .eq("status", "closed")
        .order("created_at", desc=True)
        .limit(1)
        .execute()
    )
    if not closed.data:
        raise HTTPException(404, detail="Nenhum ciclo fechado encontrado")
    cycle = closed.data[0]
    db.table("performance_cycle_reopens").insert({
        "cycle_id": cycle["id"],
        "justification": body.justification.strip(),
        "reopened_by": current_user["username"],
    }).execute()
    db.table("performance_cycles").update({"status": "open"}).eq("id", cycle["id"]).execute()
    _cache_invalidate_prefix("cycles")
    _cache_invalidate_prefix("dashboard:")
    log_action("cycle", cycle["id"], "reopen", {"status": "closed"}, {"status": "open"}, current_user["username"], request)
    return {"ok": True, "status": "open", "is_open": True}


@router.post("/cycle/send-tokens")
def send_tokens_current_cycle(
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    from services.email import send_evaluation_batch_email

    db = get_supabase()
    s = get_settings()
    frontend_url = s.allowed_origins.split(",")[0].strip().rstrip("/")

    cycle = _get_current_cycle(db)
    if not cycle or cycle["status"] != "open":
        raise HTTPException(400, detail="O ciclo precisa estar aberto para enviar tokens")

    # Buscar colaboradores a serem avaliados: L2 e L3 (Gerentes não são avaliados)
    employees_to_eval = (
        db.table("performance_employees")
        .select("id,name,cargo,email,has_corporate_email,hierarchy_level,manager_id,branch_id,company_id")
        .in_("hierarchy_level", [2, 3])
        .eq("active", True)
        .execute()
        .data
    )

    # Coletar manager_ids únicos e buscar dados dos gestores em batch
    manager_ids = {e["manager_id"] for e in employees_to_eval if e.get("manager_id")}
    managers_map: dict[str, dict] = {}
    if manager_ids:
        mgrs = (
            db.table("performance_employees")
            .select("id,name,email,has_corporate_email,hierarchy_level")
            .in_("id", list(manager_ids))
            .execute()
            .data
        )
        managers_map = {m["id"]: m for m in mgrs}

    # Buscar dados de filial/empresa em batch
    branch_ids = {e["branch_id"] for e in employees_to_eval if e.get("branch_id")}
    company_ids = {e["company_id"] for e in employees_to_eval if e.get("company_id")}
    branches_map: dict[str, str] = {}
    companies_map: dict[str, str] = {}
    if branch_ids:
        brs = db.table("performance_branches").select("id,name").in_("id", list(branch_ids)).execute().data
        branches_map = {b["id"]: b["name"] for b in brs}
    if company_ids:
        cos = db.table("performance_companies").select("id,name").in_("id", list(company_ids)).execute().data
        companies_map = {c["id"]: c["name"] for c in cos}

    sent_emails = no_email_count = tokens_created = 0

    # Pre-fetch reviews e tokens para eliminar N+1 queries
    _emp_ids_eval = [e["id"] for e in employees_to_eval]
    _prefetch_reviews = db.table("performance_reviews").select("employee_id,status").eq(
        "cycle_id", cycle["id"]
    ).eq("is_self_evaluation", False).in_("employee_id", _emp_ids_eval).execute().data or []
    _review_status_map: dict = {r["employee_id"]: r["status"] for r in _prefetch_reviews}

    _prefetch_tokens = db.table("performance_evaluation_tokens").select("employee_id,token").eq(
        "cycle_id", cycle["id"]
    ).eq("is_used", False).is_("invalidated_at", "null").in_(
        "employee_id", _emp_ids_eval
    ).execute().data or []
    _token_map: dict = {r["employee_id"]: r["token"] for r in _prefetch_tokens}

    # Agrupar colaboradores pendentes por gestor para envio de email único por gestor
    # manager_id → list[{name, cargo, company_name, branch_name, token}]
    manager_batches: dict[str, list[dict]] = {}

    for emp in employees_to_eval:
        manager_id = emp.get("manager_id")
        if not manager_id:
            no_email_count += 1
            continue

        mgr = managers_map.get(manager_id)
        if not mgr:
            no_email_count += 1
            continue

        # Bloquear se colaborador já tem avaliação concluída/calibrada/com ciência
        if _review_status_map.get(emp["id"]) in ("completed", "calibrated", "acknowledged"):
            no_email_count += 1
            continue

        # Verificar ou criar token
        token_value = _token_map.get(emp["id"])
        if token_value is None:
            token_value = str(uuid_mod.uuid4())
            tok_res = db.table("performance_evaluation_tokens").insert({
                "cycle_id": cycle["id"],
                "evaluator_id": manager_id,
                "employee_id": emp["id"],
                "token": token_value,
                "is_used": False,
                "resend_count": 0,
            }).execute()
            if not tok_res.data:
                continue
            tokens_created += 1

        if not (mgr.get("has_corporate_email") and mgr.get("email")):
            no_email_count += 1
            continue

        branch_name = branches_map.get(emp.get("branch_id", ""), "")
        company_name = companies_map.get(emp.get("company_id", ""), "")
        manager_batches.setdefault(manager_id, []).append({
            "name": emp["name"],
            "cargo": emp.get("cargo", ""),
            "company_name": company_name,
            "branch_name": branch_name,
            "token": token_value,
        })

    # Enviar um e-mail por gestor com todos os colaboradores pendentes
    for manager_id, emp_list in manager_batches.items():
        mgr = managers_map[manager_id]
        ok = send_evaluation_batch_email(
            evaluator_name=mgr["name"],
            evaluator_email=mgr["email"],
            employees=emp_list,
            cycle_name=cycle["name"],
            frontend_url=frontend_url,
        )
        if ok:
            sent_emails += 1

    log_action("cycle", cycle["id"], "send_tokens", None,
               {"sent_emails": sent_emails, "no_email_count": no_email_count, "tokens_created": tokens_created},
               current_user["username"], request)
    return {"sent_emails": sent_emails, "no_email_count": no_email_count, "tokens_created": tokens_created}


@router.get("/cycle/tokens")
def get_cycle_tokens(_: Annotated[dict, Depends(require_role(*_RH_ADMIN))]) -> list[dict]:
    db = get_supabase()
    cycle = _get_current_cycle(db)
    if not cycle:
        return []
    tokens = (
        db.table("performance_evaluation_tokens")
        .select("*")
        .eq("cycle_id", cycle["id"])
        .order("created_at", desc=True)
        .execute()
        .data
    )
    if not tokens:
        return []

    # Resolve evaluator and employee names in batch
    all_ids = set()
    for t in tokens:
        if t.get("evaluator_id"):
            all_ids.add(t["evaluator_id"])
        if t.get("employee_id"):
            all_ids.add(t["employee_id"])
    emp_map: dict[str, dict] = {}
    if all_ids:
        emps = db.table("performance_employees").select("id,name,email,has_corporate_email").in_("id", list(all_ids)).execute().data
        emp_map = {e["id"]: e for e in emps}

    result = []
    for t in tokens:
        evaluator = emp_map.get(t.get("evaluator_id", ""), {})
        employee = emp_map.get(t.get("employee_id", ""), {})
        result.append({
            "id": t["id"],
            "evaluator_id": t.get("evaluator_id"),
            "evaluator_name": evaluator.get("name", ""),
            "employee_id": t.get("employee_id"),
            "employee_name": employee.get("name", ""),
            "status": "completed" if t["is_used"] else ("invalidated" if t.get("invalidated_at") else "pending"),
            "sent_at": t.get("created_at"),
            "resend_count": t.get("resend_count", 0),
            "has_email": evaluator.get("has_corporate_email", False),
        })
    return result


@router.post("/cycle/tokens/{token_id}/resend")
def resend_cycle_token(
    token_id: str,
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    from services.email import send_evaluation_token_email

    db = get_supabase()
    s = get_settings()
    frontend_url = s.allowed_origins.split(",")[0].strip().rstrip("/")

    cycle = _get_current_cycle(db)
    if not cycle or cycle["status"] != "open":
        raise HTTPException(400, detail="O ciclo precisa estar aberto para reenviar tokens")

    tok = db.table("performance_evaluation_tokens").select("*").eq("id", token_id).eq("cycle_id", cycle["id"]).execute()
    if not tok.data:
        raise HTTPException(404, detail="Token não encontrado")
    t = tok.data[0]
    if t["is_used"]:
        raise HTTPException(400, detail="Token já utilizado — avaliação já submetida pelo gestor")
    # Bloquear reenvio se colaborador já tem avaliação concluída
    if t.get("employee_id"):
        completed = (
            db.table("performance_reviews")
            .select("status")
            .eq("cycle_id", cycle["id"])
            .eq("employee_id", t["employee_id"])
            .execute()
        )
        if completed.data and completed.data[0].get("status") in ("completed", "calibrated", "acknowledged"):
            raise HTTPException(
                400,
                detail="A avaliação deste colaborador já foi submetida. "
                       "Não é possível reenviar. Para criar uma nova avaliação, use a opção específica do RH.",
            )

    ev = db.table("performance_employees").select("*, performance_branches(name), performance_companies(name)").eq("id", t["evaluator_id"]).execute()
    if not ev.data:
        raise HTTPException(404, detail="Avaliador não encontrado")
    ev_data = ev.data[0]
    if not ev_data.get("has_corporate_email") or not ev_data.get("email"):
        raise HTTPException(400, detail="Avaliador não possui e-mail corporativo")

    employee_name = ""
    employee_cargo = ""
    if t.get("employee_id"):
        emp = db.table("performance_employees").select("name,cargo").eq("id", t["employee_id"]).execute()
        if emp.data:
            employee_name = emp.data[0]["name"]
            employee_cargo = emp.data[0].get("cargo", "")

    branch_name = (ev_data.get("performance_branches") or {}).get("name", "")
    company_name = (ev_data.get("performance_companies") or {}).get("name", "")

    ok = send_evaluation_token_email(
        evaluator_name=ev_data["name"], evaluator_email=ev_data["email"],
        employee_name=employee_name, employee_cargo=employee_cargo,
        company_name=company_name, branch_name=branch_name,
        cycle_name=cycle["name"], token=t["token"], frontend_url=frontend_url,
    )
    if ok:
        new_count = (t.get("resend_count") or 0) + 1
        db.table("performance_evaluation_tokens").update({
            "resend_count": new_count,
            "last_resent_at": datetime.now(tz=timezone.utc).isoformat(),
        }).eq("id", token_id).execute()

    log_action("token", token_id, "resend", None, {"evaluator_id": t["evaluator_id"]}, current_user["username"], request)
    return {"ok": ok}


@router.post("/cycle/send-self-evaluation-tokens")
def send_self_evaluation_tokens(
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    from services.email import send_self_evaluation_email

    db = get_supabase()
    s = get_settings()
    frontend_url = s.allowed_origins.split(",")[0].strip().rstrip("/")

    cycle = _get_current_cycle(db)
    if not cycle:
        raise HTTPException(400, detail="Nenhum ciclo ativo encontrado")
    if cycle["status"] != "open":
        raise HTTPException(400, detail="O ciclo precisa estar aberto para enviar auto-avaliações")

    employees = db.table("performance_employees").select("*").eq("active", True).execute().data
    if not employees:
        raise HTTPException(404, detail="Nenhum colaborador ativo encontrado")

    tokens_created = 0
    sent_emails = 0
    no_email_count = 0

    # Pre-fetch tokens e empresas para eliminar N+1 queries
    _emp_ids_self = [e["id"] for e in employees]
    _self_tok_rows = db.table("performance_self_evaluation_tokens").select(
        "employee_id,id,token,is_used,resend_count"
    ).eq("cycle_id", cycle["id"]).is_("invalidated_at", "null").in_(
        "employee_id", _emp_ids_self
    ).execute().data or []
    _self_tok_map: dict = {r["employee_id"]: r for r in _self_tok_rows}

    _co_ids_self = list({e["company_id"] for e in employees if e.get("company_id")})
    _companies_map_self: dict = {}
    if _co_ids_self:
        for _co in db.table("performance_companies").select("id,name").in_("id", _co_ids_self).execute().data or []:
            _companies_map_self[_co["id"]] = _co["name"]

    for emp in employees:
        # Verificar se já tem token de auto-avaliação (via pre-fetch)
        _tok_data = _self_tok_map.get(emp["id"])
        existing_tok = type("_R", (), {"data": [_tok_data] if _tok_data else []})()

        if existing_tok.data and existing_tok.data[0]["is_used"]:
            continue  # auto-avaliação já concluída

        if existing_tok.data:
            tok_val = existing_tok.data[0]["token"]
            tok_id  = existing_tok.data[0]["id"]
        else:
            import uuid as _u
            tok_val = str(_u.uuid4())
            tok_res = db.table("performance_self_evaluation_tokens").insert({
                "token": tok_val,
                "cycle_id": cycle["id"],
                "employee_id": emp["id"],
            }).execute()
            if not tok_res.data:
                continue
            tok_id = tok_res.data[0]["id"]
            tokens_created += 1

        channel_ok = False
        if emp.get("has_corporate_email") and emp.get("email"):
            company_name = _companies_map_self.get(emp.get("company_id", ""), "")
            channel_ok = send_self_evaluation_email(
                employee_name=emp["name"],
                employee_email=emp["email"],
                cycle_name=cycle["name"],
                token=tok_val,
                frontend_url=frontend_url,
                company_name=company_name,
            )
            if channel_ok:
                sent_emails += 1
        elif emp.get("whatsapp_phone"):
            from services.whatsapp import send_self_evaluation_whatsapp
            channel_ok = send_self_evaluation_whatsapp(
                employee_name=emp["name"],
                phone=emp["whatsapp_phone"],
                token=tok_val,
                cycle_name=cycle["name"],
                frontend_url=frontend_url,
            )
            if channel_ok:
                sent_emails += 1
        else:
            no_email_count += 1

        if channel_ok:
            db.table("performance_self_evaluation_tokens").update({
                "sent_at": datetime.now(tz=timezone.utc).isoformat(),
                "resend_count": (existing_tok.data[0].get("resend_count", 0) + 1) if existing_tok.data else 1,
            }).eq("id", tok_id).execute()

    _cache_invalidate_prefix("dashboard:")
    log_action("self_eval_tokens", cycle["id"], "send", None,
               {"sent": sent_emails, "no_email": no_email_count, "created": tokens_created},
               current_user["username"], request)
    return {
        "ok": True,
        "sent_emails": sent_emails,
        "no_email_count": no_email_count,
        "tokens_created": tokens_created,
    }


@router.get("/cycle/self-evaluation-tokens")
def get_self_evaluation_tokens(_: Annotated[dict, Depends(require_role(*_RH_ADMIN))]) -> list[dict]:
    db = get_supabase()
    cycle = _get_current_cycle(db)
    if not cycle:
        return []
    tokens = (
        db.table("performance_self_evaluation_tokens")
        .select("*")
        .eq("cycle_id", cycle["id"])
        .order("created_at", desc=True)
        .execute()
        .data
    )
    if not tokens:
        return []

    all_emp_ids = list({t["employee_id"] for t in tokens if t.get("employee_id")})
    emp_map: dict[str, dict] = {}
    if all_emp_ids:
        emps = db.table("performance_employees").select("id,name,email,has_corporate_email").in_("id", all_emp_ids).execute().data
        emp_map = {e["id"]: e for e in emps}

    result = []
    for t in tokens:
        emp = emp_map.get(t.get("employee_id", ""), {})
        result.append({
            "id": t["id"],
            "employee_id": t.get("employee_id"),
            "employee_name": emp.get("name", ""),
            "status": "completed" if t["is_used"] else ("invalidated" if t.get("invalidated_at") else "pending"),
            "sent_at": t.get("sent_at"),
            "resend_count": t.get("resend_count", 0),
            "has_email": emp.get("has_corporate_email", False),
        })
    return result


@router.post("/cycle/self-evaluation-tokens/{token_id}/resend")
def resend_self_eval_token(
    token_id: str,
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    from services.email import send_self_evaluation_email

    db = get_supabase()
    s = get_settings()
    frontend_url = s.allowed_origins.split(",")[0].strip().rstrip("/")

    cycle = _get_current_cycle(db)
    if not cycle or cycle["status"] != "open":
        raise HTTPException(400, detail="O ciclo precisa estar aberto para reenviar tokens de auto-avaliação")

    tok = db.table("performance_self_evaluation_tokens").select("*").eq("id", token_id).eq("cycle_id", cycle["id"]).execute()
    if not tok.data:
        raise HTTPException(404, detail="Token de auto-avaliação não encontrado")
    t = tok.data[0]

    if t.get("is_used"):
        raise HTTPException(400, detail="Auto-avaliação já concluída — não é possível reenviar")

    emp_res = db.table("performance_employees").select("*, performance_companies(name)").eq("id", t["employee_id"]).execute()
    if not emp_res.data:
        raise HTTPException(404, detail="Colaborador não encontrado")
    emp = emp_res.data[0]

    if not emp.get("has_corporate_email") or not emp.get("email"):
        raise HTTPException(400, detail="Colaborador não possui e-mail corporativo cadastrado")

    company_name = (emp.get("performance_companies") or {}).get("name", "")

    ok = send_self_evaluation_email(
        employee_name=emp["name"],
        employee_email=emp["email"],
        cycle_name=cycle["name"],
        token=t["token"],
        frontend_url=frontend_url,
        company_name=company_name,
    )
    if ok:
        new_count = (t.get("resend_count") or 0) + 1
        db.table("performance_self_evaluation_tokens").update({
            "resend_count": new_count,
            "sent_at": datetime.now(tz=timezone.utc).isoformat(),
        }).eq("id", token_id).execute()

    log_action("self_eval_token", token_id, "resend", None, {"employee_id": t["employee_id"]}, current_user["username"], request)
    return {"ok": ok}


@router.post("/cycle/self-evaluation-tokens/send-for-employee")
def send_self_eval_for_employee(
    body: dict,
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    """Cria (se não existe) e envia token de auto-avaliação para um colaborador específico."""
    from services.email import send_self_evaluation_email
    import uuid as _u

    employee_id = body.get("employee_id")
    if not employee_id:
        raise HTTPException(400, detail="employee_id é obrigatório")

    db = get_supabase()
    s = get_settings()
    frontend_url = s.allowed_origins.split(",")[0].strip().rstrip("/")

    cycle = _get_current_cycle(db)
    if not cycle or cycle["status"] != "open":
        raise HTTPException(400, detail="O ciclo precisa estar aberto")

    emp_res = db.table("performance_employees").select("*, performance_companies(name)").eq("id", employee_id).eq("active", True).execute()
    if not emp_res.data:
        raise HTTPException(404, detail="Colaborador não encontrado ou inativo")
    emp = emp_res.data[0]

    if not emp.get("has_corporate_email") or not emp.get("email"):
        raise HTTPException(400, detail="Colaborador não possui e-mail corporativo cadastrado")

    # Busca token existente ou cria novo
    tok_res = db.table("performance_self_evaluation_tokens").select("id,token,is_used,resend_count").eq("cycle_id", cycle["id"]).eq("employee_id", employee_id).is_("invalidated_at", "null").execute()

    if tok_res.data and tok_res.data[0]["is_used"]:
        raise HTTPException(400, detail="Auto-avaliação já concluída")

    if tok_res.data:
        tok_id  = tok_res.data[0]["id"]
        tok_val = tok_res.data[0]["token"]
        resend_count = (tok_res.data[0].get("resend_count") or 0) + 1
    else:
        tok_val = str(_u.uuid4())
        new_tok = db.table("performance_self_evaluation_tokens").insert({
            "token": tok_val, "cycle_id": cycle["id"], "employee_id": employee_id,
        }).execute()
        if not new_tok.data:
            raise HTTPException(500, detail="Erro ao criar token de auto-avaliação")
        tok_id = new_tok.data[0]["id"]
        resend_count = 1

    company_name = (emp.get("performance_companies") or {}).get("name", "")
    ok = send_self_evaluation_email(
        employee_name=emp["name"],
        employee_email=emp["email"],
        cycle_name=cycle["name"],
        token=tok_val,
        frontend_url=frontend_url,
        company_name=company_name,
    )
    if ok:
        db.table("performance_self_evaluation_tokens").update({
            "resend_count": resend_count,
            "sent_at": datetime.now(tz=timezone.utc).isoformat(),
        }).eq("id", tok_id).execute()

    log_action("self_eval_token", tok_id, "send_for_employee", None,
               {"employee_id": employee_id, "ok": ok}, current_user["username"], request)
    return {"ok": ok, "token_id": tok_id}


@router.get("/cycle/reopen-history")
def get_reopen_history(_: Annotated[dict, Depends(require_role(*_RH_ADMIN))]) -> list[dict]:
    db = get_supabase()
    cycle = _get_current_cycle(db)
    if not cycle:
        return []
    history = (
        db.table("performance_cycle_reopens")
        .select("*")
        .eq("cycle_id", cycle["id"])
        .order("created_at", desc=True)
        .execute()
        .data
    )
    return [
        {"created_at": h.get("created_at"), "user_name": h.get("reopened_by"), "justification": h.get("justification")}
        for h in history
    ]


# ── Evaluations management ─────────────────────────────────────────────────────

@router.get("/evaluations/export")
def export_evaluations(
    _: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
    status_filter: str | None = None,
    company_id: str | None = None,
):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    db = get_supabase()
    cycle = _get_current_cycle(db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Avaliações"

    header_fill = PatternFill("solid", fgColor="00694E")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Colaborador", "Cargo", "Nível", "Avaliador", "Nota Final", "Status", "Auto-Aval.", "Enviado em"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    STATUS_LABEL = {"pending": "Pendente", "completed": "Avaliado", "calibrated": "Calibrado",
                    "acknowledged": "Ciência Dada"}

    if cycle:
        # Usa list_evaluations para ter dados completos (inclui todos L2+L3)
        from routes.admin import list_evaluations as _le  # noqa — import circular evitado com alias
        # Chama diretamente a lógica de list_evaluations
        reviews_raw = (
            db.table("performance_reviews").select("*")
            .eq("cycle_id", cycle["id"]).eq("is_self_evaluation", False).execute().data
        )
        base_emps = (
            db.table("performance_employees")
            .select("id,name,cargo,company_id,has_corporate_email,hierarchy_level,perfil")
            .in_("hierarchy_level", [2, 3]).eq("active", True).execute().data
        )
        emp_map2 = {e["id"]: e for e in base_emps}
        rev_by_emp = {r["employee_id"]: r for r in reviews_raw if r.get("employee_id")}
        ev_ids = {r.get("evaluator_id") for r in reviews_raw if r.get("evaluator_id")}
        if ev_ids:
            mgrs = db.table("performance_employees").select("id,name").in_("id", list(ev_ids)).execute().data
            for m in mgrs: emp_map2[m["id"]] = emp_map2.get(m["id"]) or m

        self_done = {
            se["employee_id"] for se in db.table("performance_reviews").select("employee_id")
            .eq("cycle_id", cycle["id"]).eq("is_self_evaluation", True).eq("status", "completed").execute().data
        }

        for emp in sorted(base_emps, key=lambda x: x.get("name", "")):
            if company_id and emp.get("company_id") != company_id:
                continue
            r = rev_by_emp.get(emp["id"])
            if status_filter and (r.get("status") if r else "pending") != status_filter:
                continue
            ev = emp_map2.get((r.get("evaluator_id") if r else None) or "", {})
            perfil = emp.get("perfil") or ""
            lvl = {"gerente": "Gerente", "coordenador_supervisor": "Coord./Supervisor",
                   "administrativo": "Administrativo", "operacional": "Operacional"}.get(perfil, "")
            ws.append([
                emp.get("name", ""),
                emp.get("cargo", ""),
                lvl,
                ev.get("name", ""),
                float(r["final_score"]) if r and r.get("final_score") is not None else "",
                STATUS_LABEL.get(r.get("status", "pending") if r else "pending", "Pendente"),
                "Concluída" if emp["id"] in self_done else "Pendente",
                (r.get("submitted_at") or "")[:19].replace("T", " ") if r else "",
            ])
            for cell in ws[ws.max_row]:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

    for i, col in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(14, len(col) + 6)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=avaliacoes.xlsx"},
    )


@router.get("/evaluations")
def list_evaluations(
    _: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
    status: str | None = None,
    company_id: str | None = None,
    search: str | None = None,
) -> list[dict]:
    db = get_supabase()
    cycle = _get_current_cycle(db)
    if not cycle:
        return []

    # ── Base: TODOS os ativos L2+L3 do ciclo (aparecem mesmo sem token/review) ───
    base_emps_raw = (
        db.table("performance_employees")
        .select("id,name,cargo,company_id,has_corporate_email,manager_id,hierarchy_level")
        .in_("hierarchy_level", [2, 3])
        .eq("active", True)
        .execute()
        .data
    )
    emp_map: dict[str, dict] = {e["id"]: e for e in base_emps_raw}
    all_emp_ids: set[str] = set(emp_map.keys())

    if not all_emp_ids:
        return []

    # ── Reviews existentes (gestor, não auto-avaliação) ─────────────────────────
    reviews_raw = (
        db.table("performance_reviews")
        .select("*")
        .eq("cycle_id", cycle["id"])
        .eq("is_self_evaluation", False)
        .execute()
        .data
    )
    review_by_emp: dict[str, dict] = {r["employee_id"]: r for r in reviews_raw if r.get("employee_id")}

    # ── Tokens de avaliação válidos por colaborador ──────────────────────────────
    tokens_raw = (
        db.table("performance_evaluation_tokens")
        .select("id,employee_id,evaluator_id,token,is_used,invalidated_at,resend_count")
        .eq("cycle_id", cycle["id"])
        .execute()
        .data
    )
    token_by_emp: dict[str, dict] = {}
    for t in tokens_raw:
        eid = t.get("employee_id")
        if not eid or t.get("invalidated_at"):
            continue
        token_by_emp[eid] = t

    # ── Carregar dados de avaliadores (gestores) ──────────────────────────────────
    evaluator_ids: set[str] = set()
    for r in reviews_raw:
        if r.get("evaluator_id"):
            evaluator_ids.add(r["evaluator_id"])
    for t in tokens_raw:
        if t.get("evaluator_id"):
            evaluator_ids.add(t["evaluator_id"])
    # Inclui manager_id de cada colaborador como fallback para mostrar gestor esperado
    for e in base_emps_raw:
        if e.get("manager_id"):
            evaluator_ids.add(e["manager_id"])

    if evaluator_ids:
        mgr_emps = db.table("performance_employees").select("id,name").in_("id", list(evaluator_ids)).execute().data
        for m in mgr_emps:
            emp_map[m["id"]] = emp_map.get(m["id"]) or m

    # ── Acks e calibrações ────────────────────────────────────────────────────────
    review_ids = [r["id"] for r in reviews_raw]
    ack_map: dict[str, dict] = {}
    calib_map: dict[str, dict] = {}
    if review_ids:
        acks = (
            db.table("performance_review_acknowledgments")
            .select("review_id,acknowledged_at,acknowledged_via")
            .in_("review_id", review_ids)
            .order("acknowledged_at", desc=True)
            .execute()
            .data
        )
        for a in acks:
            ack_map.setdefault(a["review_id"], a)

        calibs = (
            db.table("performance_calibrations")
            .select("review_id,calibrated_at")
            .in_("review_id", review_ids)
            .order("calibrated_at", desc=True)
            .execute()
            .data
        )
        for c in calibs:
            calib_map.setdefault(c["review_id"], c)

    # ── Self-eval status ──────────────────────────────────────────────────────────
    self_eval_reviews = (
        db.table("performance_reviews")
        .select("employee_id,status")
        .eq("cycle_id", cycle["id"])
        .eq("is_self_evaluation", True)
        .execute()
        .data
    )
    self_eval_map: dict[str, str] = {se["employee_id"]: se.get("status", "pending") for se in self_eval_reviews if se.get("employee_id")}

    self_eval_tokens_raw = (
        db.table("performance_self_evaluation_tokens")
        .select("employee_id,is_used")
        .eq("cycle_id", cycle["id"])
        .execute()
        .data
    )
    self_eval_token_map: dict[str, bool] = {
        t["employee_id"]: t["is_used"] for t in self_eval_tokens_raw if t.get("employee_id")
    }

    # ── Montar resultado ──────────────────────────────────────────────────────────
    result = []
    for emp_id in all_emp_ids:
        emp = emp_map.get(emp_id, {})
        if company_id and emp.get("company_id") != company_id:
            continue
        emp_name = emp.get("name", "")
        if search and search.lower() not in emp_name.lower():
            continue

        r = review_by_emp.get(emp_id)
        tok = token_by_emp.get(emp_id)

        # Status da avaliação do gestor — independente de calibragem/ciência
        if r:
            rid = r["id"]
            # "calibrated" também conta como submetida — só a calibração do RH altera
            # esse valor na review; a ciência não grava status na tabela de reviews.
            manager_status = "completed" if r.get("status") in ("completed", "calibrated") else "pending"
        else:
            rid = None
            manager_status = "pending"

        calib_info = calib_map.get(rid, {}) if rid else {}
        ack_info = ack_map.get(rid, {}) if rid else {}
        is_calibrated = bool(rid) and rid in calib_map
        is_acknowledged = bool(rid) and rid in ack_map

        # Filtro `status`: calibragem/ciência têm valor próprio; senão filtra pelo status do gestor
        if status:
            if status == "calibrated" and not is_calibrated:
                continue
            elif status == "acknowledged" and not is_acknowledged:
                continue
            elif status in ("pending", "completed") and manager_status != status:
                continue

        # Self-eval status
        if self_eval_map.get(emp_id) == "completed" or self_eval_token_map.get(emp_id):
            self_eval_status = "completed"
        elif emp_id in self_eval_token_map:
            self_eval_status = "pending"
        else:
            self_eval_status = "not_sent"

        evaluator_id = (
            (r.get("evaluator_id") if r else None)
            or (tok.get("evaluator_id") if tok else None)
            or emp_map.get(emp_id, {}).get("manager_id")
        )
        evaluator = emp_map.get(evaluator_id or "", {})
        result.append({
            "id": rid,
            "employee_name": emp_name,
            "evaluator_name": evaluator.get("name", ""),
            "final_score": r.get("final_score") if r else None,
            "status": manager_status,
            "submitted_at": r.get("submitted_at") if r else None,
            "employee_id": emp_id,
            "evaluator_id": evaluator_id,
            "has_email": emp.get("has_corporate_email", False),
            "observations": r.get("observations") if r else None,
            "self_eval_status": self_eval_status,
            "calibrated": is_calibrated,
            "calibrated_at": calib_info.get("calibrated_at"),
            "acknowledged": is_acknowledged,
            "acknowledged_at": ack_info.get("acknowledged_at"),
            "acknowledged_via": ack_info.get("acknowledged_via"),
        })

    return sorted(result, key=lambda x: (x["employee_name"] or ""))


# ── Helpers de montagem de indicadores (gestor + auto-avaliação) ──────────────

def _build_manager_indicator_map(db, review_id: str) -> dict[str, dict]:
    """Scores do gestor por indicador, com metadata (nome, descrição, nível)."""
    scores_raw = (
        db.table("performance_indicator_scores")
        .select("indicator_id,score,justification,performance_indicators(id,name,description,hierarchy_level)")
        .eq("review_id", review_id)
        .execute()
        .data
    )
    out: dict[str, dict] = {}
    for s in scores_raw:
        ind = s.get("performance_indicators") or {}
        iid = s["indicator_id"]
        out[iid] = {
            "id": iid,
            "name": ind.get("name", "") if isinstance(ind, dict) else "",
            "description": ind.get("description", "") if isinstance(ind, dict) else "",
            "hierarchy_level": ind.get("hierarchy_level") if isinstance(ind, dict) else None,
            "manager_score": float(s["score"]),
            "manager_justification": s.get("justification") or "",
        }
    return out


def _build_self_indicator_map(db, self_review_id: str | None) -> dict[str, dict]:
    """Scores da auto-avaliação por indicador. Usado também quando não há review do gestor."""
    if not self_review_id:
        return {}
    scores_raw = (
        db.table("performance_indicator_scores")
        .select("indicator_id,score,performance_indicators(id,name,description,hierarchy_level)")
        .eq("review_id", self_review_id)
        .execute()
        .data
    )
    out: dict[str, dict] = {}
    for s in scores_raw:
        ind = s.get("performance_indicators") or {}
        iid = s["indicator_id"]
        out[iid] = {
            "id": iid,
            "name": ind.get("name", "") if isinstance(ind, dict) else "",
            "description": ind.get("description", "") if isinstance(ind, dict) else "",
            "hierarchy_level": ind.get("hierarchy_level") if isinstance(ind, dict) else None,
            "self_score": float(s["score"]),
        }
    return out


def _merge_indicator_matrix(manager_map: dict, self_map: dict) -> list[dict]:
    """União dos indicadores presentes no gestor OU na auto-avaliação (nunca só a interseção) —
    corrige o caso em que só a auto-avaliação existe e a lista de indicadores ficava vazia."""
    all_ids = set(manager_map) | set(self_map)
    result = []
    for iid in all_ids:
        m, s = manager_map.get(iid, {}), self_map.get(iid, {})
        manager_score = m.get("manager_score")
        self_score = s.get("self_score")
        entry = {
            "id": iid,
            "name": m.get("name") or s.get("name") or "",
            "description": m.get("description") or s.get("description") or "",
            "hierarchy_level": m.get("hierarchy_level") if m else s.get("hierarchy_level"),
            "manager_score": manager_score,
            "manager_justification": m.get("manager_justification", ""),
            "current_score": manager_score,  # sobrescrito por calibração, quando aplicável
            "self_score": self_score,
        }
        if manager_score and self_score:
            lo, hi = min(manager_score, self_score), max(manager_score, self_score)
            adh = round((lo / hi) * 100, 1) if hi else None
            entry["adherence_index"] = adh
            entry["adherence_label"] = (
                "alinhado" if adh >= 80 else "atencao" if adh >= 60 else "desalinhamento"
            ) if adh is not None else None
        else:
            entry["adherence_index"] = None
            entry["adherence_label"] = None
        result.append(entry)
    return sorted(result, key=lambda x: x["name"] or "")


def _calibration_history_for_review(db, review_id: str) -> list[dict]:
    calibs = (
        db.table("performance_calibrations")
        .select("id,calibrated_by,calibrated_at,notes,original_score,calibrated_score")
        .eq("review_id", review_id)
        .order("calibrated_at", desc=True)
        .execute()
        .data
    )
    calib_ids = [c["id"] for c in calibs]
    calib_items_map: dict[str, list] = {}
    if calib_ids:
        items_raw = (
            db.table("performance_calibration_items")
            .select("*")
            .in_("calibration_id", calib_ids)
            .order("created_at")
            .execute()
            .data
        )
        for item in items_raw:
            calib_items_map.setdefault(item["calibration_id"], []).append(item)
    return [{**c, "items": calib_items_map.get(c["id"], [])} for c in calibs]


# ── Detalhe da avaliação (para modal de calibração) ───────────────────────────

@router.get("/evaluations/{review_id}/detail")
def get_evaluation_detail(
    review_id: str,
    _: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    """Retorna avaliação do gestor + auto-avaliação + histórico de calibrações."""
    try:
        uuid_mod.UUID(review_id)
    except ValueError:
        raise HTTPException(400, detail="review_id inválido")

    db = get_supabase()

    rev = db.table("performance_reviews").select("*").eq("id", review_id).execute()
    if not rev.data:
        raise HTTPException(404, detail="Avaliação não encontrada")
    r = rev.data[0]

    # Colaborador e avaliador
    emp = db.table("performance_employees").select("id,name,cargo,hierarchy_level").eq("id", r["employee_id"]).execute()
    emp_data = emp.data[0] if emp.data else {}
    evaluator = db.table("performance_employees").select("id,name").eq("id", r["evaluator_id"]).execute()
    ev_data = evaluator.data[0] if evaluator.data else {}

    # Auto-avaliação do mesmo colaborador no mesmo ciclo
    self_rev = (
        db.table("performance_reviews")
        .select("id,final_score,observations")
        .eq("cycle_id", r["cycle_id"])
        .eq("employee_id", r["employee_id"])
        .eq("is_self_evaluation", True)
        .eq("status", "completed")
        .execute()
        .data
    )
    self_rev_data = self_rev[0] if self_rev else None

    manager_map = _build_manager_indicator_map(db, review_id)
    self_map = _build_self_indicator_map(db, self_rev_data["id"] if self_rev_data else None)
    indicators = _merge_indicator_matrix(manager_map, self_map)

    # Aderência geral (notas finais)
    overall_adherence_index = None
    overall_adherence_label = None
    mgr_final = r.get("final_score")
    self_final = self_rev_data.get("final_score") if self_rev_data else None
    if mgr_final and self_final:
        lo, hi = min(float(mgr_final), float(self_final)), max(float(mgr_final), float(self_final))
        overall_adherence_index = round((lo / hi) * 100, 1) if hi else None
        if overall_adherence_index is not None:
            overall_adherence_label = (
                "alinhado" if overall_adherence_index >= 80
                else "atencao" if overall_adherence_index >= 60
                else "desalinhamento"
            )

    return {
        "review": r,
        "employee": emp_data,
        "evaluator": ev_data,
        "indicators": indicators,
        "self_eval": self_rev_data,
        "calibration_history": _calibration_history_for_review(db, review_id),
        "observations": r.get("observations"),
        "self_observations": self_rev_data.get("observations") if self_rev_data else None,
        "overall_adherence_index": overall_adherence_index,
        "overall_adherence_label": overall_adherence_label,
    }


@router.get("/evaluations/detail")
def get_evaluation_detail_by_employee(
    employee_id: str,
    _: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    """Retorna avaliação do gestor (se existir) + auto-avaliação (se existir) do colaborador
    no ciclo aberto atual. Não exige que a review do gestor exista — cobre o caso em que só a
    auto-avaliação foi preenchida, que o endpoint baseado em review_id não conseguia mostrar."""
    try:
        uuid_mod.UUID(employee_id)
    except ValueError:
        raise HTTPException(400, detail="employee_id inválido")

    db = get_supabase()
    cycle = _get_current_cycle(db)
    if not cycle:
        raise HTTPException(404, detail="Nenhum ciclo aberto encontrado")

    mgr_rev = (
        db.table("performance_reviews").select("*")
        .eq("cycle_id", cycle["id"]).eq("employee_id", employee_id)
        .eq("is_self_evaluation", False).execute()
    )
    r = mgr_rev.data[0] if mgr_rev.data else None

    self_rev = (
        db.table("performance_reviews").select("id,final_score,observations")
        .eq("cycle_id", cycle["id"]).eq("employee_id", employee_id)
        .eq("is_self_evaluation", True).eq("status", "completed").execute()
    )
    self_rev_data = self_rev.data[0] if self_rev.data else None

    if not r and not self_rev_data:
        raise HTTPException(404, detail="Não há avaliações realizadas para este colaborador neste ciclo.")

    emp = db.table("performance_employees").select("id,name,cargo,hierarchy_level,manager_id").eq("id", employee_id).execute()
    emp_data = emp.data[0] if emp.data else {}

    evaluator_id = (r.get("evaluator_id") if r else None) or emp_data.get("manager_id")
    ev_data: dict = {}
    if evaluator_id:
        evaluator = db.table("performance_employees").select("id,name").eq("id", evaluator_id).execute()
        ev_data = evaluator.data[0] if evaluator.data else {}

    manager_map = _build_manager_indicator_map(db, r["id"]) if r else {}
    self_map = _build_self_indicator_map(db, self_rev_data["id"] if self_rev_data else None)
    indicators = _merge_indicator_matrix(manager_map, self_map)

    overall_adherence_index = None
    overall_adherence_label = None
    mgr_final = r.get("final_score") if r else None
    self_final = self_rev_data.get("final_score") if self_rev_data else None
    if mgr_final and self_final:
        lo, hi = min(float(mgr_final), float(self_final)), max(float(mgr_final), float(self_final))
        overall_adherence_index = round((lo / hi) * 100, 1) if hi else None
        if overall_adherence_index is not None:
            overall_adherence_label = (
                "alinhado" if overall_adherence_index >= 80
                else "atencao" if overall_adherence_index >= 60
                else "desalinhamento"
            )

    return {
        "review": r,
        "employee": emp_data,
        "evaluator": ev_data,
        "indicators": indicators,
        "self_eval": self_rev_data,
        "calibration_history": _calibration_history_for_review(db, r["id"]) if r else [],
        "observations": r.get("observations") if r else None,
        "self_observations": self_rev_data.get("observations") if self_rev_data else None,
        "overall_adherence_index": overall_adherence_index,
        "overall_adherence_label": overall_adherence_label,
    }


# ── Calibração v2 (por indicador) ─────────────────────────────────────────────

class CalibrationItem(BaseModel):
    indicator_id: str
    new_score: float
    justification: str

class CalibrateBodyV2(BaseModel):
    items: list[CalibrationItem]   # apenas os indicadores que o RH alterou
    notes: str | None = None       # observações gerais da calibração (opcional)


@router.post("/evaluations/{review_id}/calibrate", status_code=status.HTTP_201_CREATED)
def calibrate_evaluation(
    review_id: str,
    body: CalibrateBodyV2,
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    if not body.items:
        raise HTTPException(400, detail="Informe ao menos um indicador a ser calibrado")
    for item in body.items:
        if not (1 <= item.new_score <= 5):
            raise HTTPException(400, detail=f"Nota deve ser entre 1 e 5 (indicador {item.indicator_id})")
        if not item.justification.strip():
            raise HTTPException(400, detail="Justificativa é obrigatória para cada indicador alterado")

    db = get_supabase()
    rev = db.table("performance_reviews").select("final_score,cycle_id,status,employee_id").eq("id", review_id).execute()
    if not rev.data:
        raise HTTPException(404, detail="Avaliação não encontrada")

    rev_data = rev.data[0]
    rev_status = rev_data.get("status", "pending")
    if rev_status not in ("completed", "calibrated", "acknowledged"):
        raise HTTPException(400, detail="Não é possível calibrar uma avaliação pendente.")

    cycle_id = rev_data.get("cycle_id")
    cycle = db.table("performance_cycles").select("status").eq("id", cycle_id).execute()
    if not cycle.data or cycle.data[0]["status"] != "open":
        raise HTTPException(400, detail="Calibração só pode ser feita com o ciclo aberto")

    try:
        # Buscar scores atuais
        current_scores_raw = (
            db.table("performance_indicator_scores")
            .select("indicator_id,score,performance_indicators(name)")
            .eq("review_id", review_id)
            .execute()
            .data
        )
        current_map: dict[str, dict] = {}
        for s in current_scores_raw:
            ind = s.get("performance_indicators") or {}
            current_map[s["indicator_id"]] = {
                "score": float(s["score"]),
                "name": ind.get("name", s["indicator_id"]) if isinstance(ind, dict) else s["indicator_id"],
            }

        original_final = rev_data.get("final_score")

        # Criar registro de calibração (sessão)
        calib_rec = db.table("performance_calibrations").insert({
            "cycle_id": cycle_id,
            "review_id": review_id,
            "original_score": original_final,
            "calibrated_score": original_final,  # será atualizado abaixo
            "justification": (body.notes or "Calibração por indicador").strip(),
            "calibrated_by": current_user["username"],
            "notes": body.notes,
        }).execute()
        calib_id = calib_rec.data[0]["id"] if calib_rec.data else None

        # Processar cada item
        calib_items_payload = []
        for item in body.items:
            old = current_map.get(item.indicator_id, {})
            old_score = old.get("score", item.new_score)
            ind_name = old.get("name", item.indicator_id)

            # Atualizar score no indicator_scores
            db.table("performance_indicator_scores").update({
                "score": item.new_score,
            }).eq("review_id", review_id).eq("indicator_id", item.indicator_id).execute()

            # Registrar item de auditoria
            calib_items_payload.append({
                "calibration_id": calib_id,
                "indicator_id": item.indicator_id,
                "indicator_name": ind_name,
                "old_score": old_score,
                "new_score": item.new_score,
                "justification": item.justification.strip(),
            })

        if calib_items_payload and calib_id:
            db.table("performance_calibration_items").insert(calib_items_payload).execute()

        # Recalcular final_score como média de todos os scores atuais
        updated_scores = (
            db.table("performance_indicator_scores")
            .select("score")
            .eq("review_id", review_id)
            .execute()
            .data
        )
        new_final = round(sum(float(s["score"]) for s in updated_scores) / len(updated_scores), 2) if updated_scores else original_final

        # Atualizar review e calibration com novo score final
        db.table("performance_reviews").update({
            "final_score": new_final, "status": "calibrated", "updated_at": "now()",
        }).eq("id", review_id).execute()
        if calib_id:
            db.table("performance_calibrations").update({"calibrated_score": new_final}).eq("id", calib_id).execute()
    except Exception:
        _logger.exception("Falha ao calibrar review_id=%s", review_id)
        raise

    log_action("review", review_id, "calibrate",
               {"final_score": original_final},
               {"final_score": new_final, "calibrated_by": current_user["username"],
                "changed_indicators": [i.indicator_id for i in body.items]},
               current_user["username"], request)
    return {"ok": True, "new_final_score": new_final, "calibration_id": calib_id}


# ── Nova Avaliação (Override RH) ───────────────────────────────────────────────

class NewEvaluationBody(BaseModel):
    justification: str


@router.post("/employees/{employee_id}/new-evaluation", status_code=status.HTTP_201_CREATED)
def create_new_evaluation(
    employee_id: str,
    body: NewEvaluationBody,
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    """RH cria uma nova avaliação para um colaborador que já foi avaliado.
    Invalida tokens anteriores, registra o override na tabela de auditoria e cria novo token."""
    if not body.justification.strip():
        raise HTTPException(400, detail="Justificativa é obrigatória")

    db = get_supabase()
    cycle = _get_current_cycle(db)
    if not cycle or cycle["status"] != "open":
        raise HTTPException(400, detail="O ciclo precisa estar aberto para criar nova avaliação")

    emp_res = db.table("performance_employees").select("*").eq("id", employee_id).execute()
    if not emp_res.data:
        raise HTTPException(404, detail="Colaborador não encontrado")
    emp_data = emp_res.data[0]

    manager_id = emp_data.get("manager_id")
    if not manager_id:
        raise HTTPException(400, detail="Colaborador não possui gestor direto definido. Vincule um gestor antes de criar nova avaliação.")

    # Registro da avaliação anterior (se existir)
    prev_review_res = (
        db.table("performance_reviews")
        .select("id,final_score,status")
        .eq("cycle_id", cycle["id"])
        .eq("employee_id", employee_id)
        .execute()
    )
    previous_review_id = None
    previous_score = None
    if prev_review_res.data:
        previous_review_id = prev_review_res.data[0]["id"]
        previous_score = prev_review_res.data[0].get("final_score")

    # Invalidar tokens existentes para este colaborador neste ciclo
    db.table("performance_evaluation_tokens").update({
        "invalidated_at": datetime.now(tz=timezone.utc).isoformat(),
    }).eq("cycle_id", cycle["id"]).eq("employee_id", employee_id).is_("invalidated_at", "null").execute()

    # Criar novo token de avaliação
    new_token_value = str(uuid_mod.uuid4())
    tok_res = db.table("performance_evaluation_tokens").insert({
        "cycle_id": cycle["id"],
        "evaluator_id": manager_id,
        "employee_id": employee_id,
        "token": new_token_value,
        "is_used": False,
        "resend_count": 0,
    }).execute()
    if not tok_res.data:
        raise HTTPException(500, detail="Erro ao criar novo token de avaliação")
    new_token_id = tok_res.data[0]["id"]

    # Registrar override para auditoria
    db.table("performance_evaluation_overrides").insert({
        "cycle_id": cycle["id"],
        "employee_id": employee_id,
        "previous_review_id": previous_review_id,
        "previous_score": previous_score,
        "created_by": current_user["username"],
        "justification": body.justification.strip(),
        "new_token_id": new_token_id,
    }).execute()

    log_action(
        "employee", employee_id, "new_evaluation_override",
        {"previous_review_id": previous_review_id, "previous_score": previous_score},
        {"new_token_id": new_token_id, "justification": body.justification.strip()},
        current_user["username"], request,
    )

    # Enviar e-mail para o gestor (avaliador) com o novo link
    email_sent = False
    try:
        from services.email import send_evaluation_token_email
        s = get_settings()
        frontend_url = s.allowed_origins.split(",")[0].strip().rstrip("/")

        mgr_res = db.table("performance_employees").select(
            "*, performance_branches(name), performance_companies(name)"
        ).eq("id", manager_id).execute()

        if mgr_res.data:
            mgr = mgr_res.data[0]
            if mgr.get("has_corporate_email") and mgr.get("email"):
                branch_name = (mgr.get("performance_branches") or {}).get("name", "")
                company_name = (mgr.get("performance_companies") or {}).get("name", "")
                email_sent = send_evaluation_token_email(
                    evaluator_name=mgr["name"],
                    evaluator_email=mgr["email"],
                    employee_name=emp_data.get("name", ""),
                    employee_cargo=emp_data.get("cargo", ""),
                    company_name=company_name,
                    branch_name=branch_name,
                    cycle_name=cycle["name"],
                    token=new_token_value,
                    frontend_url=frontend_url,
                )
                if email_sent:
                    db.table("performance_evaluation_tokens").update({
                        "resend_count": 1,
                        "last_resent_at": datetime.now(tz=timezone.utc).isoformat(),
                    }).eq("id", new_token_id).execute()
    except Exception as exc:
        _logger.error("create_new_evaluation: erro ao enviar e-mail — %s", exc)

    return {
        "ok": True,
        "new_token_id": new_token_id,
        "new_token": new_token_value,
        "previous_review_id": previous_review_id,
        "employee_name": emp_data.get("name", ""),
        "email_sent": email_sent,
    }


# ── Nova Auto-Avaliação (Override RH) ─────────────────────────────────────────

class NewSelfEvaluationBody(BaseModel):
    justification: str


@router.post("/employees/{employee_id}/new-self-evaluation", status_code=status.HTTP_201_CREATED)
def create_new_self_evaluation(
    employee_id: str,
    body: NewSelfEvaluationBody,
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    """RH invalida a auto-avaliação anterior e cria novo token para o colaborador.
    Requer justificativa; registra auditoria com motivo e responsável."""
    if not body.justification.strip():
        raise HTTPException(400, detail="Justificativa é obrigatória")

    db = get_supabase()
    cycle = _get_current_cycle(db)
    if not cycle or cycle["status"] != "open":
        raise HTTPException(400, detail="O ciclo precisa estar aberto para criar nova auto-avaliação")

    emp_res = db.table("performance_employees").select(
        "id,name,email,has_corporate_email,company_id,performance_companies(name)"
    ).eq("id", employee_id).execute()
    if not emp_res.data:
        raise HTTPException(404, detail="Colaborador não encontrado")

    # Registro do token anterior (se existir)
    prev_tok = (
        db.table("performance_self_evaluation_tokens")
        .select("id,token,is_used")
        .eq("cycle_id", cycle["id"])
        .eq("employee_id", employee_id)
        .is_("invalidated_at", "null")
        .execute()
        .data
    )
    previous_token = prev_tok[0]["token"] if prev_tok else None
    previous_status = ("completed" if prev_tok and prev_tok[0]["is_used"] else "pending") if prev_tok else None

    # Invalidar tokens existentes
    db.table("performance_self_evaluation_tokens").update({
        "invalidated_at": datetime.now(tz=timezone.utc).isoformat(),
    }).eq("cycle_id", cycle["id"]).eq("employee_id", employee_id).is_("invalidated_at", "null").execute()

    # Criar novo token
    new_token_value = str(uuid_mod.uuid4())
    tok_res = db.table("performance_self_evaluation_tokens").insert({
        "cycle_id": cycle["id"],
        "employee_id": employee_id,
        "token": new_token_value,
        "is_used": False,
        "resend_count": 0,
    }).execute()
    if not tok_res.data:
        raise HTTPException(500, detail="Erro ao criar novo token de auto-avaliação")

    log_action(
        "self_eval_override", employee_id, "new_self_evaluation",
        old_data={"previous_token": previous_token, "previous_status": previous_status},
        new_data={"justification": body.justification.strip(), "new_token": new_token_value},
        actor=current_user["username"],
        request=request,
    )

    _cache_invalidate_prefix("dashboard:")

    # Enviar e-mail ao colaborador com o novo link de auto-avaliação
    email_sent = False
    emp_data = emp_res.data[0]
    try:
        from services.email import send_self_evaluation_email
        s = get_settings()
        frontend_url = s.allowed_origins.split(",")[0].strip().rstrip("/")
        if emp_data.get("has_corporate_email") and emp_data.get("email"):
            company_name = (emp_data.get("performance_companies") or {}).get("name", "")
            email_sent = send_self_evaluation_email(
                employee_name=emp_data["name"],
                employee_email=emp_data["email"],
                cycle_name=cycle["name"],
                token=new_token_value,
                frontend_url=frontend_url,
                company_name=company_name,
            )
            if email_sent:
                db.table("performance_self_evaluation_tokens").update({
                    "resend_count": 1,
                    "sent_at": datetime.now(tz=timezone.utc).isoformat(),
                }).eq("id", tok_res.data[0]["id"]).execute()
    except Exception as exc:
        _logger.error("create_new_self_evaluation: erro ao enviar e-mail — %s", exc)

    return {
        "ok": True,
        "new_token": new_token_value,
        "employee_name": emp_data["name"],
        "email_sent": email_sent,
    }


# ── Reset ──────────────────────────────────────────────────────────────────────

@router.post("/reset")
def reset_cycle_data(
    request: Request,
    current_user: Annotated[dict, Depends(require_role("admin"))],
) -> dict:
    db = get_supabase()
    cycle = _get_current_cycle(db)
    if not cycle:
        raise HTTPException(404, detail="Nenhum ciclo encontrado")
    cycle_id = cycle["id"]

    reviews = db.table("performance_reviews").select("id").eq("cycle_id", cycle_id).execute().data
    review_ids = [r["id"] for r in reviews]

    if review_ids:
        db.table("performance_review_acknowledgments").delete().in_("review_id", review_ids).execute()
        db.table("performance_calibrations").delete().in_("review_id", review_ids).execute()
        db.table("performance_indicator_scores").delete().in_("review_id", review_ids).execute()
        db.table("performance_acknowledgment_tokens").delete().in_("review_id", review_ids).execute()
        db.table("performance_reviews").delete().eq("cycle_id", cycle_id).execute()

    db.table("performance_evaluation_tokens").delete().eq("cycle_id", cycle_id).execute()
    db.table("performance_cycle_reopens").delete().eq("cycle_id", cycle_id).execute()
    db.table("performance_cycles").update({"status": "draft"}).eq("id", cycle_id).execute()

    log_action("cycle", cycle_id, "reset", None, {"reset_by": current_user["username"]}, current_user["username"], request)
    return {"ok": True, "cycle_id": cycle_id}


# ── Audit Log ─────────────────────────────────────────────────────────────────

@router.get("/audit-log")
def audit_log(
    _: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
    entity_type: str | None = None,
    actor: str | None = None,
    from_ts: str | None = None,
    to_ts: str | None = None,
    limit: int = 100,
    offset: int = 0,
) -> list[dict]:
    db = get_supabase()
    q = db.table("performance_audit_logs").select("*")
    if entity_type:
        q = q.eq("entity_type", entity_type)
    if actor:
        q = q.eq("actor", actor)
    if from_ts:
        q = q.gte("ts", from_ts)
    if to_ts:
        q = q.lte("ts", to_ts)
    return q.order("ts", desc=True).range(offset, offset + limit - 1).execute().data
