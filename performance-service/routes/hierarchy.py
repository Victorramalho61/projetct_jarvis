import io
import logging
import re
from typing import Annotated
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from auth import require_role
from db import get_supabase
from services.audit import log_action

router = APIRouter(prefix="/api/performance/hierarchy")
_logger = logging.getLogger(__name__)
_RH_ADMIN = ("admin", "rh")

# ── Helpers de validação ──────────────────────────────────────────────────────

def _validate_name(name: str) -> None:
    """Nome deve ter mínimo 2 palavras com >= 3 chars cada"""
    parts = name.strip().split()
    if len(parts) < 2 or any(len(p) < 3 for p in parts):
        raise HTTPException(400, detail="Nome deve ser completo, sem abreviações (mínimo 2 partes com 3+ caracteres cada)")

def _validate_matricula(matricula: str) -> None:
    """Matrícula deve conter apenas dígitos"""
    if not re.match(r'^\d+$', matricula.strip()):
        raise HTTPException(400, detail="Matrícula deve conter apenas números, sem pontos ou traços")

# ── Empresas e Filiais ────────────────────────────────────────────────────────

@router.get("/companies")
def list_companies(_: Annotated[dict, Depends(require_role(*_RH_ADMIN))]) -> list[dict]:
    db = get_supabase()
    companies = db.table("performance_companies").select("*").eq("active", True).execute().data
    branches = db.table("performance_branches").select("*").eq("active", True).execute().data
    branches_by_company: dict[str, list] = {}
    for b in branches:
        branches_by_company.setdefault(b["company_id"], []).append(b)
    for c in companies:
        c["branches"] = branches_by_company.get(c["id"], [])
    return companies

# ── Gerências ─────────────────────────────────────────────────────────────────

@router.get("/managements")
def list_managements(
    _: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
    branch_id: str | None = None,
) -> list[dict]:
    db = get_supabase()
    query = db.table("performance_managements").select("*").eq("active", True)
    if branch_id:
        query = query.eq("branch_id", branch_id)
    return query.order("name").execute().data

class ManagementCreate(BaseModel):
    branch_id: str
    name: str

@router.post("/managements", status_code=status.HTTP_201_CREATED)
def create_management(
    body: ManagementCreate,
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    if not body.name.strip():
        raise HTTPException(400, detail="Nome da gerência não pode ser vazio")
    db = get_supabase()
    result = db.table("performance_managements").insert({
        "branch_id": body.branch_id,
        "name": body.name.strip(),
        "active": True,
    }).execute()
    if not result.data:
        raise HTTPException(500, detail="Erro ao criar gerência")
    mgmt = result.data[0]
    log_action("management", mgmt["id"], "create", None, mgmt, current_user["username"], request)
    return mgmt

# ── Colaboradores ─────────────────────────────────────────────────────────────

@router.get("/employees")
def list_employees(
    current_user: Annotated[dict, Depends(require_role("admin", "rh", "gerente", "coordenador_supervisor"))],
    company_id: str | None = None,
    branch_id: str | None = None,
    management_id: str | None = None,
    hierarchy_level: int | None = None,
    active_only: bool = True,
) -> list[dict]:
    db = get_supabase()
    query = db.table("performance_employees").select("*, performance_managements(name), performance_branches(name), performance_companies(name)")

    # Gerente e coord_sup só veem sua própria equipe
    role = current_user.get("role")
    if role in ("gerente", "coordenador_supervisor"):
        # Busca o employee_id do usuário logado
        emp = db.table("performance_employees").select("id").eq("jarvis_username", current_user.get("username")).execute()
        if not emp.data:
            return []
        manager_id = emp.data[0]["id"]
        query = query.eq("manager_id", manager_id)
    else:
        if company_id:
            query = query.eq("company_id", company_id)
        if branch_id:
            query = query.eq("branch_id", branch_id)
        if management_id:
            query = query.eq("management_id", management_id)

    if hierarchy_level:
        query = query.eq("hierarchy_level", hierarchy_level)
    if active_only:
        query = query.eq("active", True)

    employees = query.order("name").execute().data
    # Enriquecer com manager_name
    all_emps = {e["id"]: e["name"] for e in db.table("performance_employees").select("id,name").execute().data}
    for e in employees:
        e["manager_name"] = all_emps.get(e.get("manager_id"), "") if e.get("manager_id") else ""
    return employees

class EmployeeCreate(BaseModel):
    name: str
    matricula: str
    email: str | None = None
    cargo: str
    has_corporate_email: bool = True
    hierarchy_level: int
    manager_id: str | None = None
    management_id: str
    branch_id: str
    company_id: str
    jarvis_username: str | None = None
    jarvis_role: str | None = None

class EmployeeUpdate(BaseModel):
    name: str | None = None
    matricula: str | None = None
    email: str | None = None
    cargo: str | None = None
    has_corporate_email: bool | None = None
    hierarchy_level: int | None = None
    manager_id: str | None = None
    management_id: str | None = None
    jarvis_username: str | None = None
    jarvis_role: str | None = None
    active: bool | None = None

@router.post("/employees", status_code=status.HTTP_201_CREATED)
def create_employee(
    body: EmployeeCreate,
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    _validate_name(body.name)
    _validate_matricula(body.matricula)
    if body.hierarchy_level not in (1, 2, 3):
        raise HTTPException(400, detail="hierarchy_level deve ser 1, 2 ou 3")
    if body.has_corporate_email and not body.email:
        raise HTTPException(400, detail="E-mail obrigatório quando has_corporate_email=True")
    db = get_supabase()
    # Checar unicidade de matrícula por empresa
    dup = db.table("performance_employees").select("id").eq("matricula", body.matricula.strip()).eq("company_id", body.company_id).execute()
    if dup.data:
        raise HTTPException(400, detail=f"Matrícula {body.matricula} já cadastrada nesta empresa")
    result = db.table("performance_employees").insert({
        "name": body.name.strip(),
        "matricula": body.matricula.strip(),
        "email": body.email,
        "cargo": body.cargo,
        "has_corporate_email": body.has_corporate_email,
        "hierarchy_level": body.hierarchy_level,
        "manager_id": body.manager_id,
        "management_id": body.management_id,
        "branch_id": body.branch_id,
        "company_id": body.company_id,
        "jarvis_username": body.jarvis_username,
        "jarvis_role": body.jarvis_role,
        "active": True,
    }).execute()
    if not result.data:
        raise HTTPException(500, detail="Erro ao cadastrar colaborador")
    emp = result.data[0]
    log_action("employee", emp["id"], "create", None, emp, current_user["username"], request)
    return emp

@router.put("/employees/{employee_id}")
def update_employee(
    employee_id: str,
    body: EmployeeUpdate,
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    db = get_supabase()
    existing = db.table("performance_employees").select("*").eq("id", employee_id).execute()
    if not existing.data:
        raise HTTPException(404, detail="Colaborador não encontrado")
    old = existing.data[0]
    updates = body.model_dump(exclude_none=True)
    if not updates:
        raise HTTPException(400, detail="Nenhum campo para atualizar")
    if "name" in updates:
        _validate_name(updates["name"])
    if "matricula" in updates:
        _validate_matricula(updates["matricula"])
        # Checar unicidade
        company_id = old.get("company_id")
        dup = db.table("performance_employees").select("id").eq("matricula", updates["matricula"]).eq("company_id", company_id).neq("id", employee_id).execute()
        if dup.data:
            raise HTTPException(400, detail=f"Matrícula {updates['matricula']} já cadastrada nesta empresa")
    updates["updated_at"] = "now()"
    result = db.table("performance_employees").update(updates).eq("id", employee_id).execute()
    if not result.data:
        raise HTTPException(500, detail="Erro ao atualizar colaborador")
    new = result.data[0]
    log_action("employee", employee_id, "update", old, new, current_user["username"], request)
    return new

@router.delete("/employees/{employee_id}")
def deactivate_employee(
    employee_id: str,
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
) -> dict:
    db = get_supabase()
    existing = db.table("performance_employees").select("*").eq("id", employee_id).execute()
    if not existing.data:
        raise HTTPException(404, detail="Colaborador não encontrado")
    result = db.table("performance_employees").update({"active": False, "updated_at": "now()"}).eq("id", employee_id).execute()
    log_action("employee", employee_id, "deactivate", existing.data[0], result.data[0], current_user["username"], request)
    return result.data[0]

# ── Template Excel ────────────────────────────────────────────────────────────

@router.get("/template-excel")
def download_template(
    _: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
):
    """Gera e retorna template Excel para importação em massa de colaboradores"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, detail="openpyxl não instalado")

    wb = openpyxl.Workbook()

    # ── Aba Listas (para os dropdowns) ─────────────────────────────────────────
    ws_lists = wb.active
    ws_lists.title = "Listas"

    ws_lists["A1"] = "Empresa"
    ws_lists["A2"] = "VTC Operadora Logística"
    ws_lists["A3"] = "Voetur Viagens"

    ws_lists["B1"] = "Filiais VTCLOG"
    for i, f in enumerate(["Brasil 21","Guarulhos","Contagem","BSB Log","Recife","Galeão - Rio de Janeiro"], 2):
        ws_lists.cell(row=i, column=2).value = f

    ws_lists["C1"] = "Filiais Voetur"
    for i, f in enumerate(["Brasília Shopping","Liberty Mall","Rio de Janeiro","São Paulo"], 2):
        ws_lists.cell(row=i, column=3).value = f

    ws_lists["D1"] = "Nível Hierárquico"
    ws_lists["D2"] = "Gerente"
    ws_lists["D3"] = "Coordenador-Supervisor"
    ws_lists["D4"] = "Operacional-Administrativo"

    ws_lists["E1"] = "Tem E-mail"
    ws_lists["E2"] = "Sim"
    ws_lists["E3"] = "Não"

    # ── Aba Instruções ─────────────────────────────────────────────────────────
    ws_inst = wb.create_sheet("Instruções")
    header_font = Font(bold=True, size=12)

    ws_inst["A1"] = "COMO PREENCHER O TEMPLATE"
    ws_inst["A1"].font = Font(bold=True, size=14)
    ws_inst["A3"] = "Exemplo de preenchimento correto:"
    ws_inst["A3"].font = header_font

    cols = ["Empresa","Filial","Gerência","Nome Completo","Matrícula","Cargo","Nível Hierárquico","E-mail Corporativo","Tem E-mail Corporativo?","Matrícula do Gestor Direto"]
    for i, c in enumerate(cols, 1):
        ws_inst.cell(row=4, column=i).value = c
        ws_inst.cell(row=4, column=i).font = Font(bold=True)

    # Exemplos
    examples = [
        ["VTC Operadora Logística","Guarulhos","Gerência de Operações","João da Silva Santos","100001","Gerente de Operações","Gerente","joao.santos@vtclog.com.br","Sim",""],
        ["VTC Operadora Logística","Guarulhos","Gerência de Operações","Maria Fernanda Costa","100002","Coordenadora de Logística","Coordenador-Supervisor","maria.costa@vtclog.com.br","Sim","100001"],
        ["VTC Operadora Logística","Guarulhos","Gerência de Operações","Carlos Eduardo Lima","100003","Supervisor de Operações","Coordenador-Supervisor","","Não","100001"],
        ["VTC Operadora Logística","Guarulhos","Gerência de Operações","Ana Paula Rodrigues","100004","Operadora Logística","Operacional-Administrativo","ana.rodrigues@vtclog.com.br","Sim","100002"],
        ["VTC Operadora Logística","Guarulhos","Gerência de Operações","Roberto Alves Pereira","100005","Auxiliar de Operações","Operacional-Administrativo","","Não","100003"],
    ]
    for r, row_data in enumerate(examples, 5):
        for c, val in enumerate(row_data, 1):
            ws_inst.cell(row=r, column=c).value = val

    ws_inst["A11"] = "OBSERVAÇÕES IMPORTANTES"
    ws_inst["A11"].font = header_font
    notes = [
        "- Nome Completo: sem abreviações. Ex: 'João da Silva Santos' (correto) / 'J. Silva' (incorreto)",
        "- Matrícula: apenas números, sem pontos ou traços. Ex: '100001' (correto) / '100.001' (incorreto)",
        "- Nível Hierárquico: use exatamente os valores da lista (Gerente / Coordenador-Supervisor / Operacional-Administrativo)",
        "- Matrícula do Gestor Direto: Gerentes (nível 1) deixam em branco. Outros devem referenciar uma matrícula existente.",
        "- Tem E-mail Corporativo?: use 'Sim' ou 'Não'. Se 'Sim', o campo E-mail Corporativo é obrigatório.",
        "- Empresa: use exatamente 'VTC Operadora Logística' ou 'Voetur Viagens'.",
    ]
    for i, note in enumerate(notes, 12):
        ws_inst.cell(row=i, column=1).value = note

    ws_inst.column_dimensions["A"].width = 80

    # ── Aba Dados (principal) ──────────────────────────────────────────────────
    ws_data = wb.create_sheet("Dados")

    # Headers
    headers = ["Empresa","Filial","Gerência","Nome Completo","Matrícula","Cargo","Nível Hierárquico","E-mail Corporativo","Tem E-mail Corporativo?","Matrícula do Gestor Direto"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font_style = Font(bold=True, color="FFFFFF")

    for i, h in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=i)
        cell.value = h
        cell.font = header_font_style
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Linha de instrução (linha 2)
    instrucoes = [
        "VTC Operadora Logística ou Voetur Viagens",
        "Nome da filial",
        "Nome da gerência",
        "Nome completo SEM abreviações",
        "Apenas números",
        "Cargo do colaborador",
        "Gerente / Coordenador-Supervisor / Operacional-Administrativo",
        "E-mail (obrigatório se Tem E-mail=Sim)",
        "Sim ou Não",
        "Matrícula do gestor (deixe vazio para Gerente)",
    ]
    inst_font = Font(italic=True, color="808080")
    for i, inst in enumerate(instrucoes, 1):
        cell = ws_data.cell(row=2, column=i)
        cell.value = inst
        cell.font = inst_font

    # Validações de lista (linhas 3 a 502)
    dv_empresa = DataValidation(type="list", formula1='"VTC Operadora Logística,Voetur Viagens"', allow_blank=False)
    dv_nivel = DataValidation(type="list", formula1='"Gerente,Coordenador-Supervisor,Operacional-Administrativo"', allow_blank=False)
    dv_email = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=False)

    ws_data.add_data_validation(dv_empresa)
    ws_data.add_data_validation(dv_nivel)
    ws_data.add_data_validation(dv_email)

    dv_empresa.sqref = "A3:A502"
    dv_nivel.sqref = "G3:G502"
    dv_email.sqref = "I3:I502"

    # Larguras das colunas
    col_widths = [28, 25, 30, 35, 12, 30, 30, 35, 22, 25]
    for i, w in enumerate(col_widths, 1):
        ws_data.column_dimensions[get_column_letter(i)].width = w

    # Ordenar abas: Dados primeiro
    wb.move_sheet("Dados", offset=-len(wb.worksheets))

    # Salvar em buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_hierarquia_avd.xlsx"},
    )

# ── Import Excel ──────────────────────────────────────────────────────────────

@router.post("/import-excel")
async def import_excel(
    request: Request,
    current_user: Annotated[dict, Depends(require_role(*_RH_ADMIN))],
    file: UploadFile = File(...),
) -> dict:
    """Importa colaboradores a partir de planilha .xlsx"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, detail="openpyxl não instalado")

    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(400, detail="Arquivo deve ser .xlsx")

    content = await file.read()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception:
        raise HTTPException(400, detail="Arquivo inválido ou corrompido")

    if "Dados" not in wb.sheetnames:
        raise HTTPException(400, detail="Aba 'Dados' não encontrada no arquivo. Use o template correto.")

    ws = wb["Dados"]

    db = get_supabase()

    # Mapa de empresas e filiais
    companies = {c["name"]: c for c in db.table("performance_companies").select("*").execute().data}
    branches_all = db.table("performance_branches").select("*").execute().data

    # Mapa de matrículas já no banco
    existing_matriculas = {
        f"{e['matricula']}_{e['company_id']}": True
        for e in db.table("performance_employees").select("matricula,company_id").execute().data
    }

    NIVEL_MAP = {
        "gerente": 1,
        "coordenador-supervisor": 2,
        "operacional-administrativo": 3,
    }

    errors = []
    rows_data = []
    file_matriculas: dict[str, str] = {}  # matricula_company_id -> linha

    # Lê a partir da linha 3 (pula header linha 1 e instrução linha 2)
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if all(v is None or str(v).strip() == "" for v in row):
            break  # linha vazia = fim dos dados

        empresa_name = str(row[0] or "").strip()
        filial_name  = str(row[1] or "").strip()
        gerencia_name= str(row[2] or "").strip()
        nome         = str(row[3] or "").strip()
        matricula    = str(row[4] or "").strip()
        cargo        = str(row[5] or "").strip()
        nivel_str    = str(row[6] or "").strip()
        email        = str(row[7] or "").strip() or None
        tem_email_str= str(row[8] or "").strip()
        gestor_mat   = str(row[9] or "").strip() or None

        row_errors = []

        # Validar empresa
        company = companies.get(empresa_name)
        if not company:
            row_errors.append({"linha": row_idx, "campo": "Empresa", "erro": f"Empresa '{empresa_name}' inválida. Use: VTC Operadora Logística ou Voetur Viagens"})

        # Validar filial
        branch = None
        if company:
            branch_matches = [b for b in branches_all if b["company_id"] == company["id"] and b["name"].lower() == filial_name.lower()]
            if not branch_matches:
                row_errors.append({"linha": row_idx, "campo": "Filial", "erro": f"Filial '{filial_name}' não encontrada para {empresa_name}"})
            else:
                branch = branch_matches[0]

        # Validar nome
        parts = nome.split()
        if len(parts) < 2 or any(len(p) < 3 for p in parts):
            row_errors.append({"linha": row_idx, "campo": "Nome Completo", "erro": "Nome deve ser completo, sem abreviações (mínimo 2 partes com 3+ caracteres)"})

        # Validar matrícula
        if not re.match(r'^\d+$', matricula):
            row_errors.append({"linha": row_idx, "campo": "Matrícula", "erro": "Matrícula deve conter apenas números, sem pontos ou traços"})
        elif company:
            key = f"{matricula}_{company['id']}"
            if key in existing_matriculas:
                row_errors.append({"linha": row_idx, "campo": "Matrícula", "erro": f"Matrícula {matricula} já cadastrada para {empresa_name}"})
            elif key in file_matriculas:
                row_errors.append({"linha": row_idx, "campo": "Matrícula", "erro": f"Matrícula {matricula} duplicada no arquivo (linha {file_matriculas[key]})"})
            else:
                file_matriculas[key] = str(row_idx)

        # Validar nível
        nivel_key = nivel_str.lower()
        hierarchy_level = NIVEL_MAP.get(nivel_key)
        if hierarchy_level is None:
            row_errors.append({"linha": row_idx, "campo": "Nível Hierárquico", "erro": f"Nível '{nivel_str}' inválido. Use: Gerente, Coordenador-Supervisor ou Operacional-Administrativo"})

        # Validar e-mail
        tem_email = tem_email_str.lower() in ("sim", "s", "yes")
        if tem_email and not email:
            row_errors.append({"linha": row_idx, "campo": "E-mail Corporativo", "erro": "E-mail obrigatório quando 'Tem E-mail Corporativo?' = Sim"})

        if row_errors:
            errors.extend(row_errors)
        else:
            rows_data.append({
                "empresa_name": empresa_name,
                "company_id": company["id"] if company else None,
                "branch_id": branch["id"] if branch else None,
                "gerencia_name": gerencia_name,
                "name": nome,
                "matricula": matricula,
                "cargo": cargo,
                "hierarchy_level": hierarchy_level,
                "email": email,
                "has_corporate_email": tem_email,
                "gestor_matricula": gestor_mat,
            })

    if errors:
        return {"errors": errors, "imported": 0}

    if not rows_data:
        return {"errors": [{"linha": 0, "campo": "Arquivo", "erro": "Nenhuma linha de dados encontrada (verifique se preencheu a partir da linha 3)"}], "imported": 0}

    # Segunda passagem: resolver gestor_matricula e criar gerências
    # Mapa: (matricula, company_id) -> employee_id (para resolver referências de gestor)
    saved_map: dict[str, str] = {}

    # Criar gerências que não existem ainda
    existing_mgmts = {
        f"{m['branch_id']}_{m['name'].lower()}": m["id"]
        for m in db.table("performance_managements").select("*").execute().data
    }

    # Inserir em ordem de hierarchy_level (primeiro os gerentes, depois coord, depois operacional)
    imported = 0
    for row in sorted(rows_data, key=lambda x: x["hierarchy_level"]):
        # Resolver ou criar gerência
        mgmt_key = f"{row['branch_id']}_{row['gerencia_name'].lower()}"
        mgmt_id = existing_mgmts.get(mgmt_key)
        if not mgmt_id:
            mgmt_res = db.table("performance_managements").insert({
                "branch_id": row["branch_id"],
                "name": row["gerencia_name"],
                "active": True,
            }).execute()
            if mgmt_res.data:
                mgmt_id = mgmt_res.data[0]["id"]
                existing_mgmts[mgmt_key] = mgmt_id

        # Resolver manager_id
        manager_id = None
        if row["gestor_matricula"]:
            mgr_key = f"{row['gestor_matricula']}_{row['company_id']}"
            # Buscar no banco primeiro
            mgr_db = db.table("performance_employees").select("id").eq("matricula", row["gestor_matricula"]).eq("company_id", row["company_id"]).execute()
            if mgr_db.data:
                manager_id = mgr_db.data[0]["id"]
            else:
                manager_id = saved_map.get(mgr_key)

        # Inserir colaborador
        emp_data = {
            "name": row["name"],
            "matricula": row["matricula"],
            "email": row["email"],
            "cargo": row["cargo"],
            "has_corporate_email": row["has_corporate_email"],
            "hierarchy_level": row["hierarchy_level"],
            "manager_id": manager_id,
            "management_id": mgmt_id,
            "branch_id": row["branch_id"],
            "company_id": row["company_id"],
            "active": True,
        }
        emp_res = db.table("performance_employees").insert(emp_data).execute()
        if emp_res.data:
            emp_id = emp_res.data[0]["id"]
            saved_map[f"{row['matricula']}_{row['company_id']}"] = emp_id
            imported += 1

    log_action("system", "import-excel", "bulk_import", None, {"imported": imported}, current_user["username"], request)
    return {"errors": [], "imported": imported}

# ── Reset Admin ───────────────────────────────────────────────────────────────

class ResetConfirm(BaseModel):
    confirm: str

@router.post("/admin/reset-all")
def reset_all(
    body: ResetConfirm,
    request: Request,
    current_user: Annotated[dict, Depends(require_role("admin"))],
) -> dict:
    if body.confirm != "RESETAR_TUDO":
        raise HTTPException(400, detail="Confirmação inválida. Digite exatamente 'RESETAR_TUDO'")

    db = get_supabase()

    # Ordem de deleção respeitando FKs (filhos primeiro)
    tables = [
        "performance_ciencia_attempts",
        "performance_review_acknowledgments",
        "performance_acknowledgment_tokens",
        "performance_indicator_scores",
        "performance_calibrations",
        "performance_evaluation_tokens",
        "performance_review_versions",
        "performance_reviews",
        "performance_cycle_reopens",
        "performance_cycles",
        "performance_employees",
        "performance_managements",
        "performance_audit_logs",
    ]

    for table in tables:
        try:
            db.table(table).delete().neq("id", "00000000-0000-0000-0000-000000000000").execute()
        except Exception as e:
            _logger.warning("Erro ao limpar tabela %s: %s", table, e)

    log_action("system", "reset", "reset_all", None, {"tables_cleared": tables}, current_user["username"], request)
    _logger.warning("BASE OPERACIONAL RESETADA por %s (IP: %s)", current_user["username"], request.client.host if request.client else "unknown")

    return {"ok": True, "message": "Base operacional limpa com sucesso. Indicadores, empresas e filiais foram preservados."}
